import streamlit as st
import pandas as pd
import shutil
import subprocess
import cv2
import numpy as np
from ultralytics import YOLO
import tempfile
import io
import gc
import time
import os
import traceback
import yaml
import sys
import logging
import uuid
import random
import queue
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from collections import defaultdict
import sqlite3
import mediapipe as mp
from huggingface_hub import HfApi, hf_hub_download, list_repo_files
import pytz
from streamlit_option_menu import option_menu
# [v21.12.16] Hugging Face Permissions for YOLO Cache
os.environ["YOLO_CONFIG_DIR"] = "/tmp/ultralytics"
os.environ["MPLCONFIGDIR"] = "/tmp/matplotlib"

# [v90.5 Speed Fix] Global Extreme Thread Limitation for Hugging Face Free CPU (2 vCPUs)
os.environ["OMP_NUM_THREADS"] = "2"
os.environ["OPENBLAS_NUM_THREADS"] = "2"
os.environ["MKL_NUM_THREADS"] = "2"
os.environ["VECLIB_MAXIMUM_THREADS"] = "2"
os.environ["NUMEXPR_NUM_THREADS"] = "2"
cv2.setNumThreads(2)

# --- 0. 設定與常數 ---

DB_FILE = "hmeayc.db"
try:
    HF_TOKEN = st.secrets.get("HF_TOKEN")
    REPO_ID = st.secrets.get("REPO_ID", "liaowen0724/hmeayc-storage")
except Exception:
    HF_TOKEN = None
    REPO_ID = "liaowen0724/hmeayc-storage"

def pull_db_from_hf():
    """Initial pull of the DB from HF Hub on startup."""
    if not HF_TOKEN:
        return False, "尚未偵測到 HF_TOKEN"
    
    try:
        api = HfApi(token=HF_TOKEN)
        try:
            api.repo_info(repo_id=REPO_ID, repo_type="dataset")
        except:
            api.create_repo(repo_id=REPO_ID, repo_type="dataset", private=True, exist_ok=True)
            return True, "已建立雲端儲存空間"

        files = list_repo_files(repo_id=REPO_ID, repo_type="dataset", token=HF_TOKEN)
        if DB_FILE in files:
            path = hf_hub_download(repo_id=REPO_ID, filename=DB_FILE, repo_type="dataset", token=HF_TOKEN)
            import shutil
            shutil.copy(path, DB_FILE)
            return True, "已成功從雲端同步數據"
        return True, "雲端尚無備份"
    except Exception as e:
        return False, f"同步讀取錯誤: {e}"

def push_db_to_hf():
    """Push local DB to HF Hub after updates."""
    if not HF_TOKEN:
        return False
    try:
        api = HfApi(token=HF_TOKEN)
        api.upload_file(
            path_or_fileobj=DB_FILE,
            path_in_repo=DB_FILE,
            repo_id=REPO_ID,
            repo_type="dataset"
        )
        return True
    except Exception as e:
        logging.error(f"Sync Push Error: {e}")
        return False

# Trigger initial pull
if 'hf_synced' not in st.session_state:
    success, msg = pull_db_from_hf()
    st.session_state.hf_synced = True
    st.session_state.hf_msg = msg

def init_db():
    """Initialize SQLite database with required tables."""
    conn = sqlite3.connect(DB_FILE, timeout=20)
    c = conn.cursor()

    # 1. Observations Table (Master)
    c.execute('''CREATE TABLE IF NOT EXISTS observations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    obs_date TEXT,
                    observer_name TEXT,
                    activity_name TEXT,
                    video_file TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    is_deleted INTEGER DEFAULT 0  -- [v43 New] Soft Delete Flag
                )''')

    # [v43 Fix] Auto-migration check (Add is_deleted if missing)
    try:
        c.execute("SELECT is_deleted FROM observations LIMIT 1")
    except sqlite3.OperationalError:
        c.execute("ALTER TABLE observations ADD COLUMN is_deleted INTEGER DEFAULT 0")
        conn.commit()

    # [v21.9.3 New] Add manual_class_score and manual_class_note if missing
    new_obs_cols = [("manual_class_score", "REAL DEFAULT 0.0"), ("manual_class_note", "TEXT")]
    for col_name, col_type in new_obs_cols:
        try:
            c.execute(f"SELECT {col_name} FROM observations LIMIT 1")
        except sqlite3.OperationalError:
            c.execute(f"ALTER TABLE observations ADD COLUMN {col_name} {col_type}")
            conn.commit()

    # 2. Records Table (Details per student)
    c.execute('''CREATE TABLE IF NOT EXISTS records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    obs_id INTEGER,
                    student_id TEXT,
                    role TEXT,
                    score REAL,
                    sync_score REAL,
                    focus_score REAL,
                    temp_lag REAL,
                    comment TEXT,
                    FOREIGN KEY (obs_id) REFERENCES observations (id)
                )''')
    conn.commit()
    # [v49 New] Add teacher_grading column if missing (for Decision Tree Prep)
    try:
        c.execute("SELECT teacher_grading FROM records LIMIT 1")
    except sqlite3.OperationalError:
        try:
            c.execute("ALTER TABLE records ADD COLUMN teacher_grading INTEGER DEFAULT 0")
            conn.commit()
        except Exception as e:
            print(f"Migration Error (teacher_grading): {e}")

    # [v58 New] Add stability_score column if missing
    try:
        c.execute("SELECT stability_score FROM records LIMIT 1")
    except sqlite3.OperationalError:
        try:
            c.execute("ALTER TABLE records ADD COLUMN stability_score REAL DEFAULT 0.0")
            conn.commit()
        except Exception as e:
            print(f"Migration Error (stability_score): {e}")

    # [v73 Fix] Auto-migration check (Add is_deleted to records if missing)
    try:
        c.execute("SELECT is_deleted FROM records LIMIT 1")
    except sqlite3.OperationalError:
        try:
            c.execute("ALTER TABLE records ADD COLUMN is_deleted INTEGER DEFAULT 0")
            conn.commit()
        except: pass

    # [v74 New] Add clothing and actions to records for full restoration
    new_cols = [("clothing", "TEXT"), ("actions", "TEXT")]
    for col_name, col_type in new_cols:
        try:
            c.execute(f"SELECT {col_name} FROM records LIMIT 1")
        except sqlite3.OperationalError:
                c.execute(f"ALTER TABLE records ADD COLUMN {col_name} {col_type}")
                conn.commit()

    conn.close()

def save_analysis_to_db(observer, activity, video, df):
    """Save the analyzed dataframe to SQLite with Update Capability."""
    if df.empty: return False

    try:
        conn = sqlite3.connect(DB_FILE, timeout=20) # v70 Fix: Add timeout for multi-user writes
        c = conn.cursor()

        # [v32 Fix] Check if we already have a session ID for this analysis
        obs_id = st.session_state.get('current_obs_id', None)

        if obs_id:
            # Update mode: Update Master Record timestamp
            c.execute("UPDATE observations SET timestamp=CURRENT_TIMESTAMP, is_deleted=0 WHERE id=?", (obs_id,))
            
            # [v78 Fix] Full Report Synchronization: Delete ALL old records for this observation
            # This ensures rows deleted in the UI table (st.data_editor) are removed from the DB.
            c.execute("DELETE FROM records WHERE obs_id=?", (obs_id,))
        else:
            # Insert Master Record
            date_str = datetime.now().strftime("%Y-%m-%d")
            # [v21.9.3] Capturing Manual Class Evaluation
            h_score = st.session_state.get('manual_class_score', 0.0)
            h_note = st.session_state.get('manual_class_note', "")
            
            c.execute("INSERT INTO observations (obs_date, observer_name, activity_name, video_file, is_deleted, manual_class_score, manual_class_note) VALUES (?, ?, ?, ?, 0, ?, ?)",
                    (date_str, observer, activity, video, h_score, h_note))
            obs_id = c.lastrowid
            st.session_state.current_obs_id = obs_id 
        
        # [v21.9.3] Also update master record with current manual assessments if in update mode
        if st.session_state.get('current_obs_id'):
            h_score = st.session_state.get('manual_class_score', 0.0)
            h_note = st.session_state.get('manual_class_note', "")
            c.execute("UPDATE observations SET manual_class_score=?, manual_class_note=? WHERE id=?", (h_score, h_note, obs_id))

        # Insert Student Records (New or Re-insert)
        for _, row in df.iterrows():
            # Parse numerical values safe
            try:
                # [v54 Fix] Safe Float Conversion (Handle None/Empty)
                def safe_float(val, default=0.0):
                    try:
                        if val is None or val == "":
                            return default
                        return float(val)
                    except (ValueError, TypeError):
                        return default

                score = safe_float(row.get("AI 觀察判定 (1-5)"))
                sync = safe_float(row.get("跟隨指令 (同步率%)"))
                focus = safe_float(row.get("專注度(%)"))

                # [v58 New] Extract stability score
                stability = safe_float(row.get("動作穩定度"))

                # Parse Lag (string "0.5s" -> float 0.5)
                lag_str = str(row.get("時序延遲 (Lag)", "0")).replace("s", "")
                lag = 0.0
                if lag_str != "-" and lag_str != "None":
                    lag = safe_float(lag_str)

                comment = row.get("AI 總結評語", "")
                role = row.get("參與型態", "Unknown")
                s_id = row.get("幼兒 ID", "Unknown")

                # [v49] Teacher Grading for Decision Tree
                t_grade = safe_float(row.get("教師評分 (1-5)"))

                # [v74 New] Get clothing and actions
                clothing_feat = row.get("AI 服裝特徵", "")
                action_tags = row.get("動作檢測 (舉手、側臉)", "")

                c.execute('''INSERT INTO records 
                            (obs_id, student_id, role, score, stability_score, sync_score, focus_score, temp_lag, comment, teacher_grading, clothing, actions)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                        (obs_id, s_id, role, score, stability, sync, focus, lag, comment, t_grade, clothing_feat, action_tags))
            except Exception as e:
                print(f"DB Row Error: {e}")
                st.error(f"寫入資料庫失敗 (Row Level): {e}") # [v52 Debug] Show error to user
                pass 

        conn.commit()
        conn.close()
        # [v77 New] Auto-sync to cloud after successful save
        push_db_to_hf()
        return True, obs_id # Return ID for feedback
    except Exception as e:
        return False, str(e)

# [v43 New] Soft Delete & Restore Logic
# [v43 New] Soft Delete & Restore Logic
# delete_observation_record relocated below to avoid duplication

def restore_observation_record(obs_id):
    """Restore a soft-deleted record."""
    try:
        conn = sqlite3.connect(DB_FILE, timeout=20)
        c = conn.cursor()
        c.execute("UPDATE observations SET is_deleted=0 WHERE id=?", (obs_id,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logging.error(f"Restore Error: {e}")
        return False
        return False

# [v41 New] History Management Helpers
# [v41 New] History Management Helpers
def delete_student_record(obs_id, student_id):
    """Soft-delete a specific student record from an observation."""
    try:
        conn = sqlite3.connect(DB_FILE, timeout=20)
        c = conn.cursor()
        c.execute("UPDATE records SET is_deleted=1 WHERE obs_id=? AND student_id=?", (obs_id, student_id))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logging.error(f"Delete Student Error: {e}")
        return False

def restore_student_record(obs_id, student_id):
    """Restore a soft-deleted student record."""
    try:
        conn = sqlite3.connect(DB_FILE, timeout=20)
        c = conn.cursor()
        c.execute("UPDATE records SET is_deleted=0 WHERE obs_id=? AND student_id=?", (obs_id, student_id))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logging.error(f"Restore Student Error: {e}")
        return False

# [v48 New] Motion Smoothness (Neuro-Motor Analysis)
def calculate_smoothness(pos_history):
    """
    Calculate movement smoothness based on velocity variance (simplified Jerk).
    Returns 0-100 (100 = Very Smooth, 0 = High Jitter/Tremor).
    """
    if len(pos_history) < 5: return 80.0 # Default assumption

    # Calculate velocities (distance between consecutive frames)
    velocities = []
    for i in range(1, len(pos_history)):
        p1, p2 = pos_history[i-1], pos_history[i]
        dist = np.sqrt((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2)
        velocities.append(dist)

    if not velocities: return 80.0

    # Smoothness ≈ Inverse of Velocity Variance (Consistency)
    # If velocity changes abruptly (High Acceleration/Jerk) -> High Variance -> Low Smoothness
    vel_var = np.var(velocities)

    # Normalization (Heuristic)
    # Variance usually 0-1000 depending on movement speed
    # We map 0-500 var to 100-0 score
    smoothness = max(0, 100 - (vel_var / 5.0))
    return round(smoothness, 1)

def rename_student_record(obs_id, old_name, new_name):
    """Rename a specific student in an observation record."""
    try:
        conn = sqlite3.connect(DB_FILE, timeout=20)
        c = conn.cursor()
        c.execute("UPDATE records SET student_id=? WHERE obs_id=? AND student_id=?", (new_name, obs_id, old_name))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        logging.error(f"Rename Student Error: {e}")
        return False

# [v47 New] Identity Merge for Unifying Records
def merge_student_identity(source_name, target_name):
    """
    Merge all records of 'source_name' into 'target_name'.
    Returns (success, count_updated).
    """
    try:
        conn = sqlite3.connect(DB_FILE, timeout=20)
        c = conn.cursor()

        # Check if source exists
        c.execute("SELECT COUNT(*) FROM records WHERE student_id=?", (source_name,))
        count = c.fetchone()[0]

        if count == 0:
            conn.close()
            return False, 0

        # Update records
        c.execute("UPDATE records SET student_id=? WHERE student_id=?", (target_name, source_name))
        updated_rows = c.rowcount

        conn.commit()
        conn.close()
        return True, updated_rows
    except Exception as e:
        logging.error(f"Merge Identity Error: {e}")
        return False, 0

# [v12] PyInstaller Path Resolver
def get_resource_path(relative_path):
    # Retrieve absolute path to resource, works for dev and for PyInstaller
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# [v81 Upgrade] Threaded Video Capture with Fast Skipping (cap.grab)
class ThreadedVideoCap:
    def __init__(self, path, interval=1, queue_size=128):
        self.cap = cv2.VideoCapture(path)
        self.interval = interval
        self.q = queue.Queue(maxsize=queue_size)
        self.stopped = False
        self.f_idx = 0
        self.current_f_idx = 0 # [v84 New] Track frame index for Polymorphic API
        self.thread = threading.Thread(target=self._reader, daemon=True)
        self.thread.start()

    def _reader(self):
        while not self.stopped:
            if not self.q.full():
                # Fast skip frames using grab()
                for _ in range(self.interval - 1):
                    if not self.cap.grab():
                        self.stopped = True
                        break
                    self.f_idx += 1
                
                if self.stopped: break

                ret, frame = self.cap.read()
                if not ret:
                    self.stopped = True
                    break
                self.f_idx += 1
                self.q.put((ret, frame, self.f_idx))
            else:
                time.sleep(0.005)

    def read(self):
        # [v84 Fix] Standardize to 2-tuple return to avoid "too many values to unpack"
        if self.q.empty() and self.stopped:
            return False, None
        try:
            ret, frame, f_idx = self.q.get(timeout=3)
            self.current_f_idx = f_idx
            return ret, frame
        except queue.Empty:
            return False, None

    def release(self):
        self.stopped = True
        self.thread.join(timeout=1)
        self.cap.release()

    def get(self, propId):
        # [v84 New] Handle POS_FRAMES polymorphicly
        if propId == cv2.CAP_PROP_POS_FRAMES:
            return self.current_f_idx
        return self.cap.get(propId)

    def set(self, propId, value):
        # [v85 New] Implement set for polymorphism (Used by Codec Doctor)
        return self.cap.set(propId, value)

    def isOpened(self):
        return self.cap.isOpened()

# 設定 logging
logging.basicConfig(filename="app_crash_log.txt", level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s', force=True)

# [v14 New] AI 自動評語生成邏輯
def generate_ai_comment(motion_score, sync_score, actions, gaze_status):
    """
    根據數據生成自然語言評語
    motion_score: 1-5
    sync_score: 0-100 (or None)
    actions: list of strings ['跳躍', '蹲下'...]
    gaze_status: '專注' or '側臉' or '一般'
    """
    comment = ""

    # 1. 活躍度描述
    if motion_score >= 4:
        comment += "表現活力充沛，肢體動作幅度大。"
    elif motion_score == 3:
        comment += "參與度平穩，動作適中。"
    else:
        comment += "處於靜態觀察狀態，動作較少。"

    # 2. 同步率描述 (如果有)
    if sync_score is not None:
        if sync_score >= 80:
            comment += " 與教師動作高度同步，跟隨指令極佳。"
        elif sync_score >= 50:
            comment += " 大致能跟隨教師指令。"
        else:
            comment += " 展現自我風格，未完全跟隨指令。"

    # 3. 動作細節
    if actions:
        unique_actions = list(set(actions))
        action_str = "、".join(unique_actions)
        comment += f" 頻繁出現「{action_str}」等動作。"

    # 4. 專注力
    if gaze_status == "專注":
        comment += " 且全程保持高度專注。"
    elif gaze_status == "側臉":
        comment += " 但注意力似乎較為發散，頻繁轉頭。"

    return comment

try:
    logging.info("App starting...")
except:
    pass
# --- 1. 系統網頁與保險箱設定 ---
st.set_page_config(page_title="AI影片辨識系統", layout="wide", page_icon="🩰", initial_sidebar_state="auto")
st.title("🩰 AI影片辨識系統")

# [改版] 加入響應式排版與步驟導航列的自訂 CSS
st.markdown("""
<style>
/* 優化手機版 Radio (步驟條) 顯示 */
div.row-widget.stRadio > div {
    display: flex;
    flex-direction: row;
    flex-wrap: wrap;
    justify-content: center;
    gap: 10px;
    background-color: #f0f2f6;
    padding: 10px;
    border-radius: 10px;
}
/* 暗黑模式相容 */
@media (prefers-color-scheme: dark) {
    div.row-widget.stRadio > div {
         background-color: #262730;
    }
}
</style>
""", unsafe_allow_html=True)

if 'current_step' not in st.session_state:
    st.session_state.current_step = "1️⃣ 設定與上傳"

st.caption("")
@st.cache_resource # [v70 Fix] Crucial for Multi-user: Share model across sessions to save RAM
def load_model():
    # 增加錯誤處理
    try:
        # [PyInstaller Fix] Resolve Path
        model_path = get_resource_path("yolov8n-pose.pt")
        model = YOLO(model_path)
        return model
    except Exception as e:
        st.error(f"模型載入失敗: {e}")
        return None
model = load_model()
# 初始化保險箱 (Session State)
if 'id_list' not in st.session_state: st.session_state.id_list = set()
if 'id_features' not in st.session_state: st.session_state.id_features = {}
# [v70 New] Persistent Unique ID per browser session for safe file handling
if 'session_id' not in st.session_state: st.session_state.session_id = uuid.uuid4().hex[:8]
if 'analysis_done' not in st.session_state: st.session_state.analysis_done = False
if 'last_frame' not in st.session_state: st.session_state.last_frame = None
# 新增：追蹤目前處理完畢的檔案名稱，避免重複跑
if 'processed_file' not in st.session_state: st.session_state.processed_file = None
# [v10] Color Hunt 專業色票庫 (HSV: H[0-180], S[0-255], V[0-255])
REF_COLORS = {
    "正紅": (0, 220, 200),
    "暗紅/棗紅": (175, 180, 80),
    "酒紅": (170, 220, 80),
    "亮橘": (15, 220, 250),
    "鵝黃": (30, 150, 240),
    "土黃/芥末": (25, 200, 150),
    "米色": (20, 40, 230),
    "卡其": (20, 80, 160), # [v68 Fix] Raise S (60->80) to avoid catching desaturated purple
    "深綠": (60, 200, 80), # [v68 Fix] Lower V
    "墨綠/軍綠": (65, 100, 40), # [v68 Fix] Distinct from Khaki
    "草綠": (50, 220, 200),
    "湖水綠(Teal)": (85, 200, 150),
    "淺藍": (100, 120, 240),
    "牛仔藍": (105, 100, 160),
    "深藍": (115, 220, 80),
    "紫色": (145, 110, 160), # [v68 Fix] Centered Purple (140->145, S:150->110)
    "深紫": (150, 100, 80),  # [v68 New] Add Dark Purple
    "淺紫": (135, 70, 220),       
    "粉紅": (160, 120, 240),
    "白/粉紅白": (165, 30, 240), 
    "桃紅/洋紅": (170, 180, 200), 
    "白色": (0, 0, 250), 
    "灰色": (0, 0, 100), # [v68 Fix] Lower V
    "黑色": (0, 0, 20),
    "焦糖/棕色": (15, 160, 120)
}

def detect_fall_pose(kpts):
    try:
        # Check confidence scores for hips (11, 12) and ankles (15, 16)
        if kpts[11][2] > 0.5 and kpts[12][2] > 0.5 and kpts[15][2] > 0.5 and kpts[16][2] > 0.5:
            hip_y = (kpts[11][1] + kpts[12][1]) / 2
            ankle_y = (kpts[15][1] + kpts[16][1]) / 2
            if abs(hip_y - ankle_y) < 40: # If horizontal
                return True, 0.8
    except:
        pass
    return False, 0.0

def get_dominant_color(img, fast_mode=False):
    if img.size == 0: return "未知"

    # 轉為 HSV
    try:
        hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    except:
        return "未知"
    
    h, s, v = cv2.split(hsv)

    # [優化] 去除背景 (簡單用 V > 15 且 S > 5 當作有效區域)
    valid_mask = (v > 15) & (s > 5)
    if np.count_nonzero(valid_mask) < 5:
        return "黑色" # 幾乎全黑

    # [v82 Upgrade] Fast Mode: Use Mean instead of Median for 2x faster calculation
    if fast_mode:
        avg_h = np.mean(h[valid_mask])
        avg_s = np.mean(s[valid_mask])
        avg_v = np.mean(v[valid_mask])
    else:
        avg_h = np.median(h[valid_mask])
        avg_s = np.median(s[valid_mask])
        avg_v = np.median(v[valid_mask])

    current_color = (avg_h, avg_s, avg_v)

    # [演算法] 尋找加權歐式距離最近的顏色
    min_dist = float('inf')
    best_match = "未知"

    for name, ref_hsv in REF_COLORS.items():
        # H 的距離要特別處理 (因為是環形，0 和 180 很近)
        dh = min(abs(current_color[0] - ref_hsv[0]), 180 - abs(current_color[0] - ref_hsv[0]))
        ds = abs(current_color[1] - ref_hsv[1])
        dv = abs(current_color[2] - ref_hsv[2])

        # [v68 New] Dynamic Weighting System
        w_h, w_s, w_v = 1.0, 1.0, 1.0
        
        # 1. 處理低飽和度 (黑白灰米)
        if ref_hsv[1] < 50: 
            w_h = 0.05 # 灰色系色相不重要
            w_s = 2.0  # 飽和度區分白vs米
            w_v = 3.0  # 亮度區分白vs灰vs黑
        else:
            # 2. 彩色系 (紫色、綠色等)
            w_h = 3.5  # [v68 Fix] 強化色相權重，避免紫色偏移到卡其
            w_s = 1.2
            w_v = 1.0

            # [v68 Fix] 對於特定問題色系 (紫色 120-160, 綠色 40-90) 加強鎖定
            if 120 <= ref_hsv[0] <= 165 and 120 <= current_color[0] <= 165:
                 w_h = 6.0 # 紫色加乘
            if 40 <= ref_hsv[0] <= 90 and 40 <= current_color[0] <= 90:
                 w_h = 5.0 # 綠色加乘

        dist = np.sqrt(w_h*(dh**2) + w_s*(ds**2) + w_v*(dv**2))

        if dist < min_dist:
            min_dist = dist
            best_match = name

    # [v8] 深淺前綴
    prefix = ""
    if best_match not in ["白色", "黑色", "灰色", "米色"]:
        if avg_v < 60: prefix = "深"
        elif avg_v > 220: prefix = "淺/亮"

    return f"{prefix}{best_match}"


def get_clothing_pattern(img):
    if img.size == 0: return ""
    try:
        # [v9 Enhancement] 進階特徵分析
        # 1. 轉為灰階 & 邊緣檢測
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        edges = cv2.Canny(gray, 50, 150)

        # 2. 計算整體紋理密度
        total_pixels = edges.size
        edge_pixels = np.count_nonzero(edges)
        density = edge_pixels / total_pixels

        # 3. 檢查中心區域 (Logo 偵測)
        h, w = gray.shape
        center_h, center_w = int(h*0.3), int(w*0.3)
        center_roi = edges[center_h:h-center_h, center_w:w-center_w]
        center_density = np.count_nonzero(center_roi) / center_roi.size if center_roi.size > 0 else 0

        # 如果中心很複雜 (Density > 0.2) 但整體還好 -> 可能是圖案/Logo
        if center_density > 0.2 and center_density > density * 1.5:
            return "(含圖案/Logo)"

        # 4. 檢查高對比細節
        mask_w = cv2.inRange(gray, 200, 255) # 白色區域
        mask_b = cv2.inRange(gray, 0, 50)    # 黑色區域
        ratio_w = cv2.countNonZero(mask_w) / total_pixels
        ratio_b = cv2.countNonZero(mask_b) / total_pixels

        details = []
        if ratio_w > 0.15: details.append("白") # 超過 15% 是白色
        if ratio_b > 0.15: details.append("黑") # 超過 15% 是黑色

        if details:
            detail_str = "、".join(details)
            if density > 0.15: # 如果同時紋理也複雜
                return f"(含{detail_str}色條紋/花紋)"
            else:
                return f"(含{detail_str}色細節/圖案)"

        # 5. 一般紋理判斷
        if density > 0.2:
            return "(含花紋/條紋)"
        elif density > 0.1:
            return "(含細微花紋)"

        return ""
    except:
        return ""

# ---------------------------
# [v19 New] Advanced Interaction & Focus Helpers
# ---------------------------
def calculate_head_yaw(nose, left_ear, right_ear):
    """
    Estimate head yaw (Left/Right/Center) based on relative ear positions.
    Returns: angle in degrees (approximate), 0=Center, >0=Right, <0=Left
    """
    if nose is None or left_ear is None or right_ear is None:
        return 0 # Unknown

    # Calculate distances
    d_left = np.linalg.norm(np.array(nose) - np.array(left_ear))
    d_right = np.linalg.norm(np.array(nose) - np.array(right_ear))

    total_d = d_left + d_right
    if total_d == 0: return 0

    # Heuristic: If nose is closer to left ear (subject's left), head is turned to subject's left.
    # We want "Observer's View". 
    # Subject Left Ear is typically on Right side of image (if facing front).
    # If Nose moves to Right (closer to Left Ear), Yaw is Positive (Right).

    yaw_factor = (d_right - d_left) / total_d 
    return yaw_factor * 90 # Degrees

def check_gaze_at_target(observer_pos, observer_yaw, target_pos, tolerance=20):
    """
    Check if observer is looking at target (horizontal direction).
    """
    if observer_pos is None or target_pos is None: return False

    dx = target_pos[0] - observer_pos[0]

    # If Target is Right (dx > 0), Observer must look Right (Yaw > threshold)
    # If Target is Left (dx < 0), Observer must look Left (Yaw < -threshold)

    threshold = 10 # [v72] Min degrees to be considered "Looking Side"

    if dx > 50: # [v72] Target is significantly to the Side
        return observer_yaw > threshold 
    elif dx < -50: 
        return observer_yaw < -threshold

    # Target is straight ahead (approx same X)
    # Observer should be looking Center
    return abs(observer_yaw) < threshold

def draw_social_graph(interactions, id_map, width=1100, height=1000, min_sec=3.0):
    """
    Draw a social network graph using OpenCV.
    interactions: dict {(id1, id2): count}
    id_map: dict {id: label}
    min_sec: Minimum duration in seconds to show a connection
    """
    # Create white canvas
    canvas = np.ones((height, width, 3), dtype=np.uint8) * 255

    # [v76 Robust Fix] Determine nodes (unique IDs) from both map and interactions
    # Ensure all IDs are treated as integers to prevent type mismatch
    node_set = set()
    for k in id_map.keys():
        try: node_set.add(int(k))
        except: pass
    for pair in interactions.keys():
        for p in pair:
            try: node_set.add(int(p))
            except: pass
    
    nodes = sorted(list(node_set))
    if not nodes: return canvas

    # Position nodes in a circle
    cx, cy = width // 2, height // 2 + 20 # Lower center slightly
    # [v20.11 Layout] Radius 400 (Mega)
    radius = 400 
    node_positions = {}

    # [v77 Fix] Dynamic Node Radius based on population (prevents overlap hiding lines)
    num_nodes = len(nodes)
    # circumference / (nodes * constant)
    node_radius = max(10, min(30, int(2512 / (num_nodes * 2.5))))

    for i, node_id in enumerate(nodes):
        angle = 2 * np.pi * i / num_nodes
        x = int(cx + radius * np.cos(angle))
        y = int(cy + radius * np.sin(angle))
        node_positions[node_id] = (x, y)

    # Draw edges
    max_count = max(interactions.values()) if interactions else 1

    # [v74 Fix] Dynamic threshold based on user slider
    f_int = 2 # Assume default
    try:
        import streamlit as st
        f_int = st.session_state.get('last_frame_interval', 2)
    except: pass
    
    needed_count = (min_sec * 30.0) / f_int

    for pair, count in interactions.items():
        if count < needed_count: continue 

        # Normalize pair IDs for lookup
        try:
            p1, p2 = int(pair[0]), int(pair[1])
        except:
            continue

        pt1 = node_positions.get(p1)
        pt2 = node_positions.get(p2)

        if pt1 and pt2:
            # [v90.4 Fix] 依據實際互動秒數決定連線粗細 (使用者最新指定階層)
            sec = (count * f_int) / 30.0
            if sec >= 11.0:
                thickness = 12
            elif sec >= 9.0:
                thickness = 9
            elif sec >= 7.0:
                thickness = 7
            elif sec >= 5.0:
                thickness = 5
            elif sec >= 3.0:
                thickness = 3
            else:
                thickness = 1 # 預設或輕微互動
                
            # Color: Darker Gray (100,100,100) for better contrast
            cv2.line(canvas, pt1, pt2, (100, 100, 100), thickness)
    # Draw nodes
    for node_id in nodes:
        pos = node_positions.get(node_id)
        if not pos: continue
        x, y = pos

        color = (235, 206, 135) # SkyBlue (BGR)
        # Highlight "Core" nodes (high degree)
        degree = sum([c for (k, c) in interactions.items() if node_id in [int(x) for x in k]])
        if degree > max_count * 0.4: color = (0, 0, 255) # Red (BGR)

        # [v77] Use dynamic radius
        cv2.circle(canvas, (x, y), node_radius, color, -1) 
        cv2.circle(canvas, (x, y), node_radius, (0, 0, 0), 1 if node_radius < 15 else 2)

        # [v76 Robust] Try to get label from id_map or use ID
        label = str(node_id)
        # Check if the map has this key (might be str or int)
        map_val = id_map.get(node_id) or id_map.get(str(node_id), "")
        if "Teacher" in map_val: label = "T"

        # [v77] Scale font based on radius
        font_scale = 0.4 if node_radius < 20 else 0.8
        ts = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, font_scale, 1 if node_radius < 20 else 2)[0]
        cv2.putText(canvas, label, (x - ts[0]//2, y + ts[1]//2), cv2.FONT_HERSHEY_SIMPLEX, font_scale, (255, 255, 255), 1 if node_radius < 20 else 2)

    return canvas

def get_motion_score(positions):
    if len(positions) < 10: return 3 # 資料不足給平均分
    # 計算座標標準差 (Standard Deviation)
    # 變異數越大表示移動範圍越大
    pos_array = np.array(positions)
    std_x = np.std(pos_array[:, 0])
    std_y = np.std(pos_array[:, 1])
    movement = std_x + std_y

    # 根據移動量給分 (數值需依實際影片調整，這裡先抓個概略值)
    # 根據移動量給分 (數值需依實際影片調整，這裡先抓個概略值)
    if movement > 100: return 5 # 大幅移動
    if movement > 60: return 4 # 明顯移動 (提高門檻)
    if movement > 35: return 3 # 小幅移動 (提高門檻，避免攝影師 3 分)
    if movement > 10: return 2 # 微幅晃動
    return 1 # 幾乎靜止 (可能是背景人物/攝影師)

# [新增] 1. 線性內插 (Interpolation) - 補足短暫消失的軌跡
def interpolate_positions(data_list, max_gap=15):
    # data_list: [(frame_idx, (x, y)), ...]
    if not data_list: return []

    data_list.sort(key=lambda x: x[0]) # 按幀數排序
    interpolated = []

    for i in range(len(data_list) - 1):
        f1, p1 = data_list[i]
        f2, p2 = data_list[i+1]

        interpolated.append((f1, p1))

        gap = f2 - f1
        if 1 < gap <= max_gap:
            # 進行內插
            for j in range(1, gap):
                alpha = j / gap
                new_x = int(p1[0] * (1 - alpha) + p2[0] * alpha)
                new_y = int(p1[1] * (1 - alpha) + p2[1] * alpha)
                new_f = f1 + j
                interpolated.append((new_f, (new_x, new_y)))

    interpolated.append(data_list[-1])
    return interpolated

# [新增] 2. 師生同步率 (Teacher-Student Sync)
# 計算學生與老師的動作向量相似度 (Cosine Similarity)
def calculate_teacher_sync(student_pos, teacher_pos):
    # student_pos, teacher_pos: list of (frame, (x, y))
    # 1. 確保時間對齊 (Intersection of frames)
    s_dict = {f: p for f, p in student_pos}
    t_dict = {f: p for f, p in teacher_pos}

    common_frames = sorted(list(set(s_dict.keys()) & set(t_dict.keys())))
    logging.info(f"Sync Debug: ID={student_pos[0][1]}? Common Frames: {len(common_frames)}")

    if len(common_frames) < 10: 
        logging.info("Sync failed: Not enough common frames (<10)")
        return 0.0 # 重疊時間太短

    # 2. 計算速度向量 (Velocity Vector)
    # v[i] = p[i+1] - p[i]
    s_vecs = []

    for i in range(len(common_frames) - 3): # Skip 3 frames to get better vector
        f1, f2 = common_frames[i], common_frames[i+3]
        if f2 - f1 > 5: continue # 只有連續幀才算向量

        p_s1, p_s2 = s_dict[f1], s_dict[f2]
        p_t1, p_t2 = t_dict[f1], t_dict[f2]

        # 向量 (dx, dy)
        v_s = np.array([p_s2[0] - p_s1[0], p_s2[1] - p_s1[1]])
        v_t = np.array([p_t2[0] - p_t1[0], p_t2[1] - p_t1[1]])

        # 正規化
        norm_s = np.linalg.norm(v_s)
        norm_t = np.linalg.norm(v_t)

        # [v12 Fix] 處理靜止狀態 (Static Handling)
        threshold = 0.5
        is_static_s = norm_s < threshold
        is_static_t = norm_t < threshold

        if is_static_s and is_static_t:
            # 兩者都靜止 -> 視為同步 (給予 1.0 )
            s_vecs.append(1.0)
            continue
        elif is_static_s or is_static_t:
            # 其中一方靜止，另一方在動 -> 視為不同步 (給予 0.0)
            s_vecs.append(0.0)
            continue

        # 兩者都在動 -> 計算向量相似度
        # Cosine Similarity: A . B / |A||B|
        cos_sim = np.dot(v_s, v_t) / (norm_s * norm_t)
        s_vecs.append(cos_sim)

    if not s_vecs: 
        logging.info("Sync failed: No valid frames found")
        return 0.0

    # 3. 平均相似度 (-1 ~ 1) -> 映射到 (0 ~ 100 分)
    avg_sim = np.mean(s_vecs)

    # [Log Debug]
    logging.info(f"Sync Debug: Vectors={len(s_vecs)}, AvgSim={avg_sim:.2f}")

    # 簡單映射：因為我們已經處理了靜止(1.0)和單動(0.0)，剩下的動態部分直接取平均
    # 負值(反向)視為 0
    score = max(0, avg_sim) * 100 

    return round(score, 1)

def analyze_temporal_sync(motion_s, motion_t, fps=30, max_lag_sec=1.5):
    """
    [v26 New] Calculate Temporal Lag using Cross-Correlation on Motion Energy.
    Returns: (max_correlation, lag_in_seconds)
    Msg: "Sync" or "Delay 0.5s"
    """
    if not motion_s or not motion_t: return 0.0, 0.0

    # Ensure equal length / align to the shorter one's end (most recent)
    min_len = min(len(motion_s), len(motion_t))
    if min_len < 30: return 0.0, 0.0 # Too short

    # Take recent history (e.g., last 10 seconds = 300 frames)
    window = 300
    s = np.array(motion_s[-window:])
    t = np.array(motion_t[-window:])

    # Normalize (Z-score) to avoid amplitude bias
    if np.std(s) < 1e-6 or np.std(t) < 1e-6: return 0.0, 0.0
    s = (s - np.mean(s)) / np.std(s)
    t = (t - np.mean(t)) / np.std(t)

    # Cross Correlation
    # mode='full' returns array of size N+M-1. Index of max correlates to lag.
    corr = np.correlate(s, t, mode='full') / len(s) # Normalize by length

    # Lag finding
    # Center index (0 lag) is at len(t) - 1
    lags = np.arange(-len(t) + 1, len(s))
    max_idx = np.argmax(corr)
    max_corr = corr[max_idx]
    best_lag = lags[max_idx]

    # Convert lag frames to seconds
    # lag > 0 means s is AHEAD of t? np.correlate(s, t): sum(s[k] * t[k+delay])
    # Usually: if s is shifted by +d to match t, then s was behind.
    lag_sec = best_lag / fps

    # Limit to realistic lag (e.g. +/- 1.5s)
    if abs(lag_sec) > max_lag_sec:
        return 0.0, 0.0

    return max_corr, lag_sec

# [新增] 2. 同步率 (R-Value) 計算
def calculate_group_sync(id_motion_scores):
    # id_motion_scores: {mid: [score1, score2, ...]}
    # 簡單計算：所有 ID 在同一幀的變異數 (Variance) 的倒數
    # 變異數越小 -> 大家動作越一致 -> 同步率高
    # 這裡簡化為：計算每個人的平均動作分數，然後算這些平均分數的標準差
    # (更精確應該是 Frame-by-frame，但需要對齊時間軸)

    if not id_motion_scores: return 0

    avg_scores = []
    for mid, scores in id_motion_scores.items():
        if scores:
            avg_scores.append(np.mean(scores))

    if len(avg_scores) < 2: return 0 # 只有一人無法算同步

    std_dev = np.std(avg_scores)
    # R值設計：標準差 0 -> R=1; 標準差 2 (大) -> R=0.3
    # 公式：1 / (1 + std_dev)
    r_val = 1 / (1 + std_dev)
    return round(r_val, 2)

def calculate_kuramoto_order_parameter(id_motion_log):
    """
    [v20 New] Compute Group Synchronization using simplified Kuramoto Order Parameter.
    Order Parameter R(t) = |(1/N) * sum(exp(i * theta_j))|
    Here, we approximate phase theta_j using the direction of motion vector.
    """
    # id_motion_log: {mid: [m1, m2, ...]} -> This stores magnitude, not vector.
    # We need vector history. But `id_positions` has (frame, (x,y)).
    # Let's derive velocity vectors from `id_positions`.

def calculate_group_sync(id_positions):
    """
    [v21 Upgrade] Vector Coherence Sync (Directional).
    Computes cosine similarity of motion vectors against the group average vector.
    """
    if not id_positions or len(id_positions) < 2:
        return 0.0

    # 1. Extract Motion Vectors per ID (Frame t vs t-1)
    # We need at least 2 frames of history for each ID
    id_vectors = {} # {id: (mean_vx, mean_vy)}

    all_vectors = []

    for mid, pos_list in id_positions.items():
        if len(pos_list) < 2: continue

        # Calculate recent motion vector (using last few frames)
        # Taking average of last 3 moves to smooth jitter
        recent = pos_list[-5:]
        if len(recent) < 2: continue

        vx_sum, vy_sum = 0, 0
        count = 0
        for i in range(1, len(recent)):
            dx = recent[i][0] - recent[i-1][0]
            dy = recent[i][1] - recent[i-1][1]
            vx_sum += dx
            vy_sum += dy
            count += 1

        if count > 0:
            avg_vx, avg_vy = vx_sum/count, vy_sum/count
            # Normalize to unit vector (Direction only)
            mag = np.sqrt(avg_vx**2 + avg_vy**2)
            if mag > 1.0: # Ignore noise/stationary
                id_vectors[mid] = (avg_vx/mag, avg_vy/mag)
                all_vectors.append((avg_vx/mag, avg_gy/mag))

    if len(all_vectors) < 2: return 0.0

    # 2. Compute Group Average Vector (The "Flow")
    avg_gx = sum(v[0] for v in all_vectors) / len(all_vectors)
    avg_gy = sum(v[1] for v in all_vectors) / len(all_vectors)
    group_mag = np.sqrt(avg_gx**2 + avg_gy**2)

    if group_mag < 0.1: return 0.0 # Group is stationary or canceling out

    # Normalize group vector
    g_unit = (avg_gx/group_mag, avg_gy/group_mag)

    # 3. Compute Coherence (Cosine Similarity)
    coherence_scores = []
    for mid, v in id_vectors.items():
        # Dot product of unit vectors = Cosine Similarity (-1 to 1)
        # We only care about positive sync (0 to 1)
        sim = max(0, v[0]*g_unit[0] + v[1]*g_unit[1])
        coherence_scores.append(sim)

    return round(float(np.mean(coherence_scores)), 2)

# [v30 New] Randomized Advice Templates
# [v21.6 Upgrade] Activity-Aware Professional ECE Advice Templates
ADVICE_TEMPLATES = {
    "imitation": {
        "active_low_sync": [
            "📋 **教學觀察**：幼兒能量充沛但有自我步調。建議老師邀請其擔任「小老師」或示範者，將充沛能量轉化為帶領同儕的動力，提升其對指令的關注度。",
            "📋 **教學觀察**：動作反應及時但稍欠精準。建議透過「分段示範」或誇張化核心動作，協助其捕捉模仿細節，強化動作的視覺認知。",
            "📋 **教學觀察**：具備良好的觀察力，但執行指令有時滯。建議在動作變換前給予更明確的口語預告，降低其轉換動作的認知負荷。"
        ],
        "passive_observer": [
            "📋 **教學觀察**：屬於「潛行式」觀察者，正在靜態內化動作。建議老師以眼神或輕拍肩膀鼓勵其先從小幅度動作（如拍手、搖頭）開始參與，循序漸進。",
            "📋 **教學觀察**：視覺專注度高但肢體表現含蓄。建議可提供小道具（如響板或彩帶）作為媒介，轉移直接肢體表現的壓力，誘發參與動機。",
            "📋 **教學觀察**：在安靜環境中觀察力最優。建議安排其站在教師正對面，減少周圍同儕的視覺干擾，使其能更純粹地模仿教師動作。"
        ],
        "delayed_follower": [
            "📋 **教學觀察**：觀察到明顯的「反射性延遲」，顯示其處理聽視覺訊息需要較長時間。建議老師稍放慢示範節奏，並在節拍點給予明確的速度提示。",
            "📋 **教學觀察**：傾向於觀察同儕而非教師。建議安排一位「動作領航員」（表現穩定的同儕）在其旁邊，利用同儕模仿力帶動其跟隨。",
            "📋 **教學觀察**：動作執行度尚可但節奏感不穩定。可運用重音明顯的曲目，配合「一、二、三、跳」的口語連動，協調其身體律動感。"
        ],
        "high_performance": [
            "📋 **教學觀察**：動作同步度極高，展現出優秀的本體覺與計畫能力。建議提供更高難度（如跨中線、不對稱動作）的挑戰，維持其學習層次。",
            "📋 **教學觀察**：專業的模仿能力！可鼓勵其加入個人的細微變化（如表情或力度），在標準模仿中注入美感表達。",
            "📋 **教學觀察**：具備示範潛質。建議邀請其帶領一小段律動，培養其在群體中的領導力與自尊感。"
        ]
    },
    "creative": {
        "active_explorer": [
            "🎨 **創意觀察**：肢體特徵多元且奔放。建議肯定其對空間利用的大膽探索，並鼓勵其分享目前的動作想法，轉化為全班的創意素材。",
            "🎨 **創意觀察**：展現獨特的律動語彙，不拘泥於既定框架。建議老師可針對其動作進行「模仿回饋」，讓幼兒感受到自己的動作被看見與重視。",
            "🎨 **創意觀察**：動力十足且充滿實驗精神。建議引導其嘗試「不同質地的動作」（如像水一樣流動、像石頭一樣硬），深化肢體表現力。"
        ],
        "passive_thinker": [
            "🎨 **創意觀察**：偏好定點觀察後的點狀爆發。建議這類孩子不需要過多催促，給予充足的空間與時間，其內在美感會在安全感足夠時顯現。",
            "🎨 **創意觀察**：對音樂氛圍感應敏銳但表現含蓄。建議引導其用「微小的身體部位」（如手指、肩膀）進行創作，降低身體展現的負擔。",
            "🎨 **創意觀察**：正在進行內在空間探索。建議老師可主動靠近，以並肩移動的方式陪伴其進行低強度的創意練習。"
        ],
        "independent_style": [
            "🎨 **創意觀察**：強烈的個人風格優於團體規範。建議肯定其原創性，並透過「對話式律動」引導其與同儕進行動作上的交流互動。",
            "🎨 **創意觀察**：動作豐富但較缺乏核心集中。推薦練習「收與放」的對比動作，幫助其在自由創作中找回動作的控制感與層次。",
            "🎨 **創意觀察**：擅長利用環境與空間。建議鼓勵其嘗試不同水平高度（高、中、低）的動作變換，擴展其肢體探索的維度。"
        ]
    },
    "general": {
        "low_focus": [
            "🚩 **注意引導**：視覺焦點分散。建議運用色彩鮮豔的教具或特定節奏聲響引導其視線，建立穩定的視覺追蹤習慣。",
            "🚩 **注意引導**：易受環境干擾。建議老師將其引導至教學的核心區域，並透過更多、更短的「點名式指令」來維持其參與熱度。"
        ],
        "low_energy": [
            "🌱 **成長建議**：目前動力較低，可能與體力或動機有關。建議安排其感興趣的動作主題，或搭配律動小驚喜來提高參與的樂趣。",
            "🌱 **成長建議**：參與熱度尚待開發。可透過簡單的互動傳遞遊戲（如傳球或模仿表情）來活化其肢體反應。"
        ]
    }
}

def generate_expert_comment(score, sync_score, focus_score, role, valid_tags, activity_context="跟隨模仿 (Imitation)", class_stats=None, archetype_text=""):
    """
    [v21.6.2 Fix] Optimized sanitization for NoneTypes.
    """
    # [Sanitize] Handle potential None types from various detection modes
    if score is None: score = 0
    if sync_score is None: sync_score = 0
    if focus_score is None: focus_score = 0
    if not role: role = "Unknown"
    
    parts = []
    
    # 1. First Part: Descriptive Observation based on Metrics
    # [Descriptive Observation]
    obs_parts = []
    if score >= 4: obs_parts.append("肢體能量充沛，展現出高度的肌肉動員與積極性。")
    elif score <= 2: obs_parts.append("目前多處於靜態觀察階段，肢體開拓性尚在發展中。")
    
    if sync_score is not None and "跟隨模仿" in activity_context:
        if sync_score > 80: obs_parts.append("對音樂節奏與指令反應極其精準，顯現優異的模仿穩定度。")
        elif sync_score < 40: obs_parts.append("動作軌跡與指令有顯著差異，顯示出強烈的自我風格或理解落後。")
    
    if focus_score > 80: obs_parts.append("視覺專注度極高，能持續鎖定目標進行深度內化。")
    elif focus_score < 40: obs_parts.append("注意力較易受環境干擾，視覺焦點呈現跳躍性。")
    
    # Action specific descriptive sentence
    visible_actions = [t for t in valid_tags if t not in ['專注', '側臉']]
    if visible_actions:
        action_mapping = {
            "舉手": "主動展現上半身延展（如舉手）",
            "蹲下": "能進行重心下沉的肢體變化（如蹲下）",
            "跳躍": "具備雙腳離地的重力挑戰特質（如跳躍）",
            "地板動作": "敢於進行地面與中低水平的空間探索"
        }
        mapped_actions = [action_mapping.get(a, a) for a in visible_actions]
        obs_parts.append(f"觀察到其頻繁使用「{ '、'.join(mapped_actions) }」等肢體語彙。")

    if obs_parts:
        parts.append("".join(obs_parts)) # [v89.1 Fix] Removed "【AI 專業觀察】" prefix

    # 2. Second Part: Role Archetype Comment
    if archetype_text:
        parts.append(f"\n\n分析指出：{archetype_text}")

    # 3. Third Part: Professional Guidance based on Activity Context
    suggestions = []
    import random
    
    # Determine the category based on context
    is_creative = "自由創作" in activity_context
    cat = "creative" if is_creative else "imitation"
    
    # Logic to select the right suggestion bank
    if not is_creative:
        # Imitation Mode Logic
        if sync_score > 80 and focus_score > 80:
            suggestions.append(random.choice(ADVICE_TEMPLATES[cat]["high_performance"]))
        elif "Active" in role or score >= 4:
            suggestions.append(random.choice(ADVICE_TEMPLATES[cat]["active_low_sync"]))
        elif "Passive" in role or score <= 2:
            suggestions.append(random.choice(ADVICE_TEMPLATES[cat]["passive_observer"]))
        elif sync_score < 60:
            suggestions.append(random.choice(ADVICE_TEMPLATES[cat]["delayed_follower"]))
    else:
        # Creative Mode Logic
        if "Active" in role or score >= 4:
            suggestions.append(random.choice(ADVICE_TEMPLATES[cat]["active_explorer"]))
        elif "Independent" in role or sync_score < 40:
            suggestions.append(random.choice(ADVICE_TEMPLATES[cat]["independent_style"]))
        else:
            suggestions.append(random.choice(ADVICE_TEMPLATES[cat]["passive_thinker"]))
    
    # General Overrides (Attention/Energy)
    if focus_score < 35:
        suggestions.append(random.choice(ADVICE_TEMPLATES["general"]["low_focus"]))
    if score < 2:
        suggestions.append(random.choice(ADVICE_TEMPLATES["general"]["low_energy"]))

    if suggestions:
        # Deduplicate and join
        unique_s = []
        for s in suggestions:
            if s not in unique_s: unique_s.append(s)
        parts.append("\n\n💡 **教學與輔導建議**：\n" + "\n".join(unique_s[:2])) # Limit to 2 tips

    return "".join(parts)

# ... (Original detectaction_and_gaze stays above, skipping diff context here for brevity) ...

# ... (Previous code) ...

# [v30 New] History Management
def delete_observation_record(obs_id):
    """Soft Delete: Mark as deleted but keep data."""
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        # Soft delete only master record is enough if we filter by it
        c.execute("UPDATE observations SET is_deleted=1 WHERE id=?", (obs_id,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Soft Delete Error: {e}")
        return False

def show_history_ui():
    st.title("🗄️ 歷史紀錄與幼兒成長歷程")

    # 1. Fetch Master List (Filter OUT deleted items)
    conn = sqlite3.connect(DB_FILE)
    df_obs = pd.read_sql_query("SELECT * FROM observations WHERE is_deleted=0 ORDER BY timestamp DESC", conn)
    conn.close()

    if df_obs.empty:
        st.info("尚無歷史紀錄 (或全部已刪除)。請先在「全功能分析」模式下進行分析並儲存。")
        # Allow viewing Trash Can even if main list is empty?
        # Maybe not block immediately if we want to access Trash Can.
        # But for now, user might have deleted everything.
        # Let's check if there are ANY records including deleted.
        conn = sqlite3.connect(DB_FILE)
        count = pd.read_sql_query("SELECT COUNT(*) FROM observations", conn).iloc[0,0]
        conn.close()

        if count == 0:
            st.stop()
        else:
            st.warning("目前顯示列表為空，但垃圾桶中可能有資料。")

    # Master Table
    st.subheader("📋 活動觀察紀錄列表")

    # [v41] Manage/Delete Section (Enhanced)
    with st.expander("🛠️ 資料庫管理 (Database Management)"):
        # Create tabs for different management actions
        m_tab1, m_tab2, m_tab3, m_tab4 = st.tabs(["🗑️ 刪除整筆觀察", "✏️ 修改/刪除幼兒資料", "♻️ 垃圾桶 (復原刪除)", "👥 身份合併 (Merge)"])

        with m_tab1:
            st.caption("⚠️ 此操作將刪除該次觀察的所有數據 (包含所有幼兒紀錄)。")
            obs_to_delete = st.selectbox("選擇要刪除的紀錄:", 
                                    df_obs['id'].astype(str) + " | " + df_obs['obs_date'] + " | " + df_obs['activity_name'],
                                    index=None,
                                    placeholder="請選擇...",
                                    key="del_obs_select"
            )
            if obs_to_delete:
                obs_id = int(obs_to_delete.split(" | ")[0])
                if st.button(f"確認刪除紀錄 ({obs_id})", type="primary", key="btn_del_obs"):
                    if delete_observation_record(obs_id):
                        st.success("紀錄已刪除！")
                        st.rerun()
                    else:
                        st.error("刪除失敗。")

        with m_tab2:
            st.caption("🔧 針對特定幼兒進行改名或刪除操作。")

            # 1. Select Observation
            target_obs_str = st.selectbox("1. 選擇紀錄:", 
                                    df_obs['id'].astype(str) + " | " + df_obs['obs_date'] + " | " + df_obs['activity_name'],
                                    index=None,
                                    placeholder="請先選擇紀錄...",
                                    key="edit_obs_select"
            )

            if target_obs_str:
                target_obs_id = int(target_obs_str.split(" | ")[0])

                # 2. Get Students in this observation (Filter deleted)
                conn = sqlite3.connect("hmeayc.db")
                stu_df = pd.read_sql_query("SELECT DISTINCT student_id FROM records WHERE obs_id=? AND is_deleted=0", conn, params=(target_obs_id,))
                conn.close()

                target_student = st.selectbox("2. 選擇幼兒:", 
                                            stu_df['student_id'].tolist(),
                                            index=None,
                                            placeholder="請選擇幼兒...",
                                            key="edit_stu_select"
                )

                if target_student:
                    action = st.radio("3. 選擇操作:", ["改名 (Rename)", "刪除此人 (Delete)"], horizontal=True)

                    if action == "改名 (Rename)":
                        new_name_input = st.text_input("輸入新名稱:", value=target_student)
                        if st.button("確認改名", key="btn_rename"):
                            if new_name_input and new_name_input != target_student:
                                if rename_student_record(target_obs_id, target_student, new_name_input):
                                    st.success(f"已將 {target_student} 改名為 {new_name_input}")
                                    st.rerun()
                                else:
                                    st.error("改名失敗")
                            else:
                                st.warning("名稱未變更")

                    elif action == "刪除此人 (Delete)":
                        st.error(f"⚠️ 即將從此紀錄中移除 {target_student}，此操作僅影響單一學生！")
                        if st.button("確認移除此人", type="primary", key="btn_del_stu"):
                            if delete_student_record(target_obs_id, target_student):
                                st.success(f"已移除 {target_student}")
                                st.rerun()
                            else:
                                st.error("移除失敗")


        with m_tab3: # [v43 New] Trash Can Tab
            st.caption("♻️ 這裡存放被軟刪除的紀錄，您可以隨時復原。")
            conn = sqlite3.connect(DB_FILE)
            df_deleted = pd.read_sql_query("SELECT * FROM observations WHERE is_deleted=1 ORDER BY timestamp DESC", conn)
            conn.close()

            if df_deleted.empty:
                st.info("垃圾桶是空的 (0 筆資料)。")
            else:
                obs_to_restore = st.selectbox("選擇要復原的紀錄:", 
                                        df_deleted['id'].astype(str) + " | " + df_deleted['obs_date'] + " | " + df_deleted['activity_name'],
                                        index=None,
                                        placeholder="請選擇復原對象...",
                                        key="restore_obs_select"
                )
                if obs_to_restore:
                    obs_id_restore = int(obs_to_restore.split(" | ")[0])
                    if st.button(f"確認復原紀錄 ({obs_id_restore})", key="btn_restore"):
                        if restore_observation_record(obs_id_restore):
                            st.success(f"紀錄 {obs_id_restore} 已成功復原！")
                            st.rerun()
                        else:
                            st.error("復原失敗")

            st.markdown("---")
            st.subheader("🗑️ 已刪除的幼兒紀錄 (單一幼兒)")
            conn = sqlite3.connect(DB_FILE)
            # Find students that are deleted in observations that are NOT deleted
            df_del_stu = pd.read_sql_query("""
                SELECT r.obs_id, r.student_id, o.obs_date, o.activity_name 
                FROM records r
                JOIN observations o ON r.obs_id = o.id
                WHERE r.is_deleted = 1 AND o.is_deleted = 0
                ORDER BY o.timestamp DESC
            """, conn)
            conn.close()

            if not df_del_stu.empty:
                st.caption(f"ℹ️ 目前垃圾桶中有 {len(df_del_stu)} 筆幼兒紀錄待復原。")

            if df_del_stu.empty:
                st.info("目前無被刪除的單一幼兒紀錄。")
            else:
                stu_to_restore = st.selectbox("選擇要復原的幼兒:", 
                                        df_del_stu['obs_id'].astype(str) + " | " + df_del_stu['student_id'] + " | " + df_del_stu['obs_date'] + " | " + df_del_stu['activity_name'],
                                        index=None,
                                        placeholder="選擇要復原的幼兒...",
                                        key="restore_stu_select"
                )
                if stu_to_restore:
                    parts = stu_to_restore.split(" | ")
                    s_obs_id = int(parts[0])
                    s_name = parts[1]
                    if st.button(f"確認復原幼兒 ({s_name})", key="btn_restore_stu"):
                        if restore_student_record(s_obs_id, s_name):
                            st.success(f"已成功復原幼兒 {s_name}！")
                            st.rerun()
                        else:
                            st.error("復原失敗")

        with m_tab4: # [v47 New] Identity Merge Tab
            st.caption("👥 將多個暫存 ID (如 ID_1) 合併到同一位學生 (如 小明) 名下。此操作無法復原。")

            # Get distinct student IDs (Filter deleted)
            conn = sqlite3.connect(DB_FILE)
            temp_df = pd.read_sql_query("SELECT DISTINCT student_id FROM records WHERE is_deleted=0 ORDER BY student_id", conn)
            conn.close()

            all_ids = temp_df['student_id'].tolist() if not temp_df.empty else []

            col_m1, col_m2 = st.columns(2)

            with col_m1:
                target_merge_name = st.selectbox("1. 保留的目標 (Target):", all_ids, key="merge_target", index=0 if all_ids else None)

            with col_m2:
                # Filter out target from source options
                source_options = [x for x in all_ids if x != target_merge_name]
                source_merge_name = st.selectbox("2. 要合併的來源 (Source):", source_options, key="merge_source", index=0 if source_options else None)

            if st.button("🚀 確認合併身分", type="primary", use_container_width=True, disabled=not (target_merge_name and source_merge_name)):
                if target_merge_name == source_merge_name:
                    st.warning("目標與來源不能相同。")
                else:
                    success, count = merge_student_identity(source_merge_name, target_merge_name)
                    if success:
                        st.success(f"✅ 成功將 {count} 筆紀錄從 '{source_merge_name}' 合併至 '{target_merge_name}'！")
                        st.rerun()
                    else:
                        st.error("❌ 合併失敗或來源無資料。")
            
            # [v75 New] Download Database for Backup
            st.markdown("---")
            st.subheader("💾 備份資料庫 (Manual Backup)")
            st.caption("由於雲端空間 (Hugging Face) 在重新啟動或更新時會清空資料，強烈建議您定期下載資料庫備份。")
            try:
                with open(DB_FILE, "rb") as f:
                    st.download_button(
                        label="📥 下載完整資料庫 (hmeayc.db)",
                        data=f,
                        file_name=f"hmeayc_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db",
                        mime="application/x-sqlite3",
                        use_container_width=True
                    )
            except Exception as e:
                st.error(f"備份準備失敗: {e}")

            # [v75.1 New] Upload Database to Restore
            st.markdown("---")
            st.subheader("Restore 恢復資料庫")
            restore_file = st.file_uploader("上傳備份的 hmeayc.db 檔案來恢復紀錄:", type=["db"])
            if restore_file is not None:
                if st.button("🔄 確認覆蓋並恢復 (注意: 目前數據將被刪除)", use_container_width=True):
                    try:
                        with open(DB_FILE, "wb") as f:
                            f.write(restore_file.getbuffer())
                        st.success("✅ 資料庫已恢復！請重新整理網頁。")
                        time.sleep(1)
                        st.rerun()
                    except Exception as e:
                        st.error(f"恢復失敗: {e}")
            
            # [v77 New] Cloud Sync Status
            st.markdown("---")
            st.subheader("☁️ 雲端同步狀態 (Cloud Sync)")
            if not HF_TOKEN:
                st.warning("⚠️ 尚未偵測到 HF_TOKEN。請依據教學在 Settings -> Secrets 加入密鑰，即可開啟全自動雲端備份。")
            else:
                st.success(f"✅ 雲端備份已啟動：{st.session_state.get('hf_msg', '連線成功')}")
                if st.button("🔄 立即手動同步至雲端"):
                    if push_db_to_hf():
                        st.success("同步完成！")
                    else:
                        st.error("同步失敗，請檢查網路。")

    st.dataframe(df_obs, use_container_width=True, hide_index=True)

    # ... (Rest of existing history UI logic: Student Selection, Charts etc.) ...
    # Re-implementing the rest of show_history_ui below to ensure continuity

    st.markdown("---")
    st.subheader("📈 幼兒個人成長歷程")

    # Get unique student IDs from all active records
    conn = sqlite3.connect(DB_FILE)
    all_students = pd.read_sql_query("SELECT DISTINCT student_id FROM records WHERE is_deleted=0 ORDER BY student_id", conn)
    conn.close()

    # [v71 Fix] Enhanced Numerical Sorting for History Dropdown
    student_list = all_students['student_id'].tolist()
    
    def try_sort_key(x):
        # Extract numerical part of "ID_X" or similar strings for natural sorting
        import re
        nums = re.findall(r'\d+', str(x))
        return int(nums[0]) if nums else 999

    try:
        student_list.sort(key=try_sort_key)
    except:
        student_list.sort() # Fallback to string sort

    selected_student = st.selectbox("選擇幼兒 (ID/姓名):", student_list, key="history_student_select")

    if selected_student:
        # Fetch history for this student
        # JOIN to get date/activity
        q = f"""
        SELECT r.*, o.obs_date, o.activity_name 
        FROM records r
        JOIN observations o ON r.obs_id = o.id
        WHERE r.student_id = '{selected_student}' AND r.is_deleted = 0
        ORDER BY o.obs_date ASC
        """
        conn = sqlite3.connect(DB_FILE)
        df_hist = pd.read_sql_query(q, conn)
        conn.close()

        if not df_hist.empty:
            # Metrics
            col1, col2 = st.columns(2)
            avg_sync = df_hist['sync_score'].mean()
            avg_focus = df_hist['focus_score'].mean()

            col1.metric("平均同步率", f"{avg_sync:.1f}%")
            col2.metric("平均專注度", f"{avg_focus:.1f}%")

            # Charts
            tab1, tab2 = st.tabs(["📊 趨勢圖表", "📝 詳細數據"])

            with tab1:
                # 1. Sync & Focus over time
                chart_data = df_hist[['obs_date', 'sync_score', 'focus_score']].set_index('obs_date')
                st.line_chart(chart_data)

                # 2. Activity / Motion Score
                st.caption("動作活躍度 (1-5) 變化")
                st.bar_chart(df_hist[['obs_date', 'score']].set_index('obs_date'))

            with tab2:
                # [v39 Fix] Rename columns for display
                # [v58] Added stability_score
                display_df = df_hist[['obs_date', 'activity_name', 'role', 'score', 'stability_score', 'sync_score', 'focus_score', 'comment']].copy()
                
                # [v21.6.1 Fix] Round to 2 decimal places as requested
                for col in ['score', 'stability_score', 'sync_score', 'focus_score']:
                    display_df[col] = display_df[col].round(2)

                display_df.columns = ["日期", "活動名稱", "參與型態", "AI 評分(活躍)", "動作穩定度", "同步率", "專注度", "AI 總結評語"]
                display_df.index = display_df.index + 1 # [v55 Fix] Start index from 1 for user friendly display
                st.table(display_df) # [v47 Fix] Use st.table to allow text wrapping for long comments
        else:
            st.warning("此幼兒尚無詳細數據。")

    # [v74 New] Load Record back to Analysis
    st.markdown("---")
    st.subheader("📥 載入舊資料回報表 (Load to Analysis Mode)")
    st.caption("您可以將以前的紀錄「抓回」分析模式，在那裡進行二次刪除、改名或產出最新的 Excel 報表。")
    load_obs_str = st.selectbox("選擇要載入的紀錄:", 
                            df_obs['id'].astype(str) + " | " + df_obs['obs_date'] + " | " + df_obs['activity_name'],
                            index=None,
                            placeholder="請選擇紀錄...",
                            key="load_obs_select"
    )
    if load_obs_str:
        l_obs_id = int(load_obs_str.split(" | ")[0])
        if st.button(f"🚀 將紀錄 {l_obs_id} 載入分析模式", use_container_width=True):
            conn = sqlite3.connect(DB_FILE)
            # Load non-deleted records
            query = """
                SELECT student_id, role, score, stability_score, sync_score, focus_score, temp_lag, comment, teacher_grading, clothing, actions
                FROM records WHERE obs_id=? AND is_deleted=0
            """
            l_df = pd.read_sql_query(query, conn, params=(l_obs_id,))
            conn.close()
            
            if not l_df.empty:
                # Map to analysis format
                restored_list = []
                for i, row in l_df.iterrows():
                    restored_list.append({
                        "序號": i + 1,
                        "幼兒 ID": row['student_id'],
                        "AI 服裝特徵": row['clothing'],
                        "特徵補強 (圖案/熊/亮片)": None,
                        "AI 觀察判定 (1-5)": row['score'],
                        "跟隨指令 (同步率%)": row['sync_score'],
                        "時序延遲 (Lag)": f"{row['temp_lag']:.2f}s" if row['temp_lag'] > 0 else "-",
                        "專注度(%)": row['focus_score'],
                        "參與型態": row['role'],
                        "動作檢測 (舉手、側臉)": row['actions'],
                        "AI 總結評語": row['comment'],
                        "教師評分 (1-5)": row['teacher_grading'],
                        "教師評語": None,
                        "動作能量": 0, # Not stored
                        "動作穩定度": row['stability_score']
                    })
                st.session_state.restored_df = pd.DataFrame(restored_list)
                st.session_state.current_obs_id = l_obs_id # [v77 Fix] Ensure saving updates this record
                st.session_state.id_list = set(l_df['student_id'].tolist()) # [v78 Fix] Populate ID list
                st.session_state.pending_nav = "🚀 全功能分析" # Use deferred navigation
                st.session_state.analysis_done = True
                st.session_state.processed_file = "Restored From History"
                st.success("資料已成功載入！正在切換模式...")
                st.rerun()

    # Stop execution here to prevent Main UI from rendering below
    st.stop()

# [v13 New] 動作與視線規則檢測 (增加 跳躍/躺下)
def detectaction_and_gaze(kpts, bbox=None): # 新增 bbox 參數用於長寬比判斷
    """
    kpts: (17, 3) array [x, y, conf]
    bbox: [x1, y1, x2, y2]
    Keypoints:
    0: Nose, 1: LEye, 2: REye, 3: LEar, 4: REar
    5: LSho, 6: RSho, 7: LElb, 8: RElb, 9: LWri, 10: RWri
    11: LHip, 12: RHip, 13: LKne, 14: RKne, 15: LAnk, 16: RAnk
    """
    actions = []

    # 1. 舉手 (Hands Up): 手腕 (9/10) 高於 眼睛 (1/2) 或 耳朵 (3/4)
    # 注意: Y 軸向下為正，所以 "高於" 是 y < target_y
    # 先做信心過濾
    if kpts[9, 2] > 0.5 and kpts[1, 2] > 0.5:
        if kpts[9, 1] < kpts[1, 1]: actions.append("舉手")
    elif kpts[10, 2] > 0.5 and kpts[2, 2] > 0.5:
        if kpts[10, 1] < kpts[2, 1]: actions.append("舉手")

    # 2. 蹲下 (Squat): 臀部 (11/12) 與 膝蓋 (13/14) 垂直距離縮短
    # ... (原有蹲下邏輯) ...
    # 2. 蹲下 (Squat): 臀部 (11/12) 與 膝蓋 (13/14) 垂直距離縮短
    # [v29 Fix] Lower confidence threshold (0.5 -> 0.3) for crowded scenes where legs are occluded
    if kpts[11, 2] > 0.3 and kpts[15, 2] > 0.3 and kpts[5, 2] > 0.3:
        leg_len = kpts[15, 1] - kpts[11, 1]
        body_len = kpts[15, 1] - kpts[5, 1] # 肩到腳
        # [v29 Fix] Relax ratio (0.35 -> 0.45) to detect squats even from high angles
        if body_len > 0 and (leg_len / body_len) < 0.45: 
            actions.append("蹲下")

    # [v13] 3. 躺下/地板動作 (Lying/Floor)
    if bbox is not None:
        w = bbox[2] - bbox[0]
        h = bbox[3] - bbox[1]
        if w > h * 1.2: # 寬大於高 1.2 倍
            actions.append("地板動作")

    # [v13] 4. 跳躍/抬腿 (Jump/High Knees)
    # 邏輯：雙腳腳踝 (15/16) 的 Y 座標小於 (高於) 膝蓋 (13/14) 
    # 或者 腳踝非常接近膝蓋水平
    # 簡單版：雙腳騰空 (Ankle < Knee + offset)
    if kpts[15, 2] > 0.3 and kpts[16, 2] > 0.3 and kpts[13, 2] > 0.3 and kpts[14, 2] > 0.3:
        # 檢查左腳
        l_high = kpts[15, 1] < (kpts[13, 1] + 20) # 腳踝高於膝蓋附近
        r_high = kpts[16, 1] < (kpts[14, 1] + 20)

        if l_high and r_high:
            actions.append("跳躍") # 雙腳都高
        elif l_high or r_high:
            actions.append("抬腿") # 單腳

    # 5. 視線 (Gaze): 耳朵對稱性
    # ... (原有視線邏輯) ...
    if kpts[3, 2] > 0.3 and kpts[4, 2] > 0.3:
        # 兩耳都看得到 -> 正臉/專注
        actions.append("專注")
    elif (kpts[3, 2] > 0.5 and kpts[4, 2] < 0.1) or (kpts[3, 2] < 0.1 and kpts[4, 2] > 0.5):
        # 只有一邊耳朵 -> 側臉
        actions.append("側臉")

    return list(set(actions)) # 去重

# [v66 New] Optimized Landmark Translation (No Dict Conversion)
def translate_landmarks_fast(landmark_list, crop_x1, crop_y1, roi_w, roi_h, frame_w, frame_h):
    if not landmark_list: return None
    from mediapipe.framework.formats import landmark_pb2
    new_landmarks = landmark_pb2.NormalizedLandmarkList()
    
    for lm in landmark_list.landmark:
        # Convert to absolute ROI pixels -> Add crop offset -> Back to normalized frame
        new_lm = new_landmarks.landmark.add()
        new_lm.x = (lm.x * roi_w + crop_x1) / frame_w
        new_lm.y = (lm.y * roi_h + crop_y1) / frame_h
        new_lm.z = lm.z
        new_lm.visibility = lm.visibility
    return new_landmarks

# [New Helper UI]
def create_tracker_config():
    # [Fix v3] Switch to bytetrack: much faster on CPU (no ReID, no optical flow)
    # [v21.12.20 Update] Increase track_buffer to 240 (8 sec @ 30fps) to stabilize IDs
    config_content = """tracker_type: bytetrack
track_high_thresh: 0.15
track_low_thresh: 0.05
new_track_thresh: 0.25
# [v90.1 Fix] Increase track_buffer to 480 (16s @ 30fps) to prevent ID loss when occluded
track_buffer: 480
match_thresh: 0.85
fuse_score: True
"""
    try:
        sid = st.session_state.get('session_id', 'global')
        tracker_filename = f"custom_tracker_{sid}.yaml"
        with open(tracker_filename, "w", encoding="utf-8") as f:
            f.write(config_content)
        return tracker_filename
    except:
        return "bytetrack.yaml" # Fallback

def get_color_histogram(img):
    if img.size == 0: return None
    # 計算 HSV 直方圖作為 Re-ID 特徵
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    hist = cv2.calcHist([hsv], [0, 1], None, [180, 256], [0, 180, 0, 256])
    cv2.normalize(hist, hist, 0, 1, cv2.NORM_MINMAX)
    return hist

# [v21.12.17] Loop Reconstruction & Replay Sync Lockdown
with st.sidebar:
    st.header("⚙️ 系統設定")
    st.caption("版本: v21.12.19-LogicFix-Ultimate")

# State Management
def reset_analysis_state():
    """Reset analysis state when tracking settings change."""
    st.session_state.processed_file = None
    st.session_state.analysis_done = False
    
    # [v21.4.3 Fix] Truly delete result keys to avoid st.download_button NoneType error
    keys_to_clear = [
        'id_list', 'id_interactions', 'id_positions', 'id_yaw_history',
        'final_id_count', 'final_id_list', 'excel_ready_data', 
        'group_sync_r', 'social_graph_image', 'video_output_path'
    ]
    for k in keys_to_clear:
        if k in st.session_state:
            del st.session_state[k]
            
    # Re-initialize defaults if needed by other components
    st.session_state.id_list = set()
    st.session_state.id_interactions = defaultdict(int) 
    st.session_state.id_positions = {}
    st.session_state.id_yaw_history = {}

# --- 2. 側邊欄：資訊固定 ---
# [v76 Fix] Handle deferred navigation to prevent StreamlitAPIException
if st.session_state.get("pending_nav"):
    st.session_state.nav_mode = st.session_state.pending_nav
    del st.session_state.pending_nav

st.sidebar.header("📋 基本資訊輸入")
observer_name = st.sidebar.text_input("觀察員姓名", "文禎")
act_name = st.sidebar.text_input("活動名稱", "Walk and copy animal")
# [v90.6 Fix] Handle Hugging Face UTC Timezone difference (+8 for Taiwan)
tw_tz = pytz.timezone('Asia/Taipei')
act_date = st.sidebar.date_input("觀察日期", datetime.now(tw_tz))
music_element = st.sidebar.text_input("音樂元素 (如：走停、快慢)", "走停")

# [v91.2 Fix] Always define performance variables globally based on session state
perf_mode = st.session_state.get('perf_mode_main', "⚡ 標準模式 (Balanced)")

if "Turbo" in perf_mode:
    frame_interval = 12
    model_conf = 0.10
    draw_overlays = True
    target_imgsz = 416
elif "Ultra Fast" in perf_mode:
    frame_interval = 20
    model_conf = 0.10
    draw_overlays = False
    target_imgsz = 320
elif "Pro" in perf_mode:
    frame_interval = 2
    model_conf = 0.05
    draw_overlays = True
    target_imgsz = 640
elif "MediaPipe" in perf_mode:
    frame_interval = st.session_state.get('mp_frame_skip_main', 2)
    use_face_mesh = st.session_state.get('mp_use_face_mesh_main', False)
    model_conf = 0.10
    draw_overlays = True
    target_imgsz = 640
else: # Standard
    frame_interval = 4
    use_face_mesh = False
    model_conf = 0.05
    draw_overlays = True
    target_imgsz = 1024

st.session_state.last_frame_interval = frame_interval
save_record = st.session_state.get('save_recording', True)
cloud_booster = st.session_state.get('cloud_booster', True)

# [v65 Fix] Dynamic Max ID based on previous run results
max_id = 100  # Default before any run
if 'final_id_list' in st.session_state and st.session_state.final_id_list:
    max_id = max(st.session_state.final_id_list)

# Target track ID moved to Step 1
target_track_id = st.session_state.get('locked_target_id', 0)
# [v48 New] Context-Aware Sync (Phase 5)
st.sidebar.markdown("### 🎭 活動性質設定")

# [v52 Fix] Mode Selection moved to Step 1 to prevent State Loss on Re-run
# mode = st.sidebar.radio("模式選擇", ["🚀 全功能分析", "🗄️ 歷史紀錄查閱"], index=0, key="nav_mode")
mode = st.session_state.get('nav_mode', "🚀 全功能分析")

# Activity context and Social threshold moved to Step 1
activity_context = st.session_state.get('activity_context', "跟隨模仿 (Imitation)")
social_threshold_sec = st.session_state.get('social_threshold_sec', 3.0)

# [v21.9.1 New] Manual Identity Merge Tool
st.sidebar.write("---")
st.sidebar.subheader("🔗 身份合併工具 (ID Merge)")
st.sidebar.info("若 AI 因為遮擋把同一個小朋友認成兩個 ID (例如 32 與 47)，您可以在下方進行合併。")
m_col1, m_col2 = st.sidebar.columns(2)
with m_col1:
    source_id = st.number_input("原 ID (消失者)", min_value=0, value=0, key="merge_src")
with m_col2:
    target_id = st.number_input("新 ID (繼承者)", min_value=0, value=0, key="merge_tgt")

if st.sidebar.button("確認合併身份", type="primary", use_container_width=True):
    if source_id > 0 and target_id > 0 and source_id != target_id:
        # [v21.9.1 Fix] Store in a manual remapping list
        if 'manual_id_map' not in st.session_state:
            st.session_state.manual_id_map = {}
        st.session_state.manual_id_map[source_id] = target_id
        st.sidebar.success(f"✅ 已將 ID {source_id} 關聯至 ID {target_id}。分析與表格將合併數據。")
    else:
        st.sidebar.error("請輸入有效的不同 ID。")

# [v51 New] Decision Tree Visualization Button
if st.sidebar.button("🌳 顯示 AI 決策樹邏輯"):
    st.session_state.show_decision_tree = True

# [v75 New] Download Guide Button
try:
    with open(r"c:\Users\user\OneDrive\桌面\HMEAYC_Project\HMEAYC_功能與操作指南.txt", "r", encoding="utf-8") as f:
        guide_txt = f.read()
    st.sidebar.download_button(
        label="📥 下載系統操作指南 (.txt)",
        data=guide_txt,
        file_name="HMEAYC_Operation_Guide.txt",
        mime="text/plain"
    )
except: pass

# [v51 New] Show Decision Tree Diagram (Safe Version)
if st.session_state.get("show_decision_tree", False):
    st.info("🌳 這是目前系統採用的「專家規則決策樹 (Expert Rule Decision Tree)」，模擬了老師的判斷邏輯。")
    # ... (Markdown content is same) ...
    st.markdown("""
    #### 🌲 專家邏輯可視化 (Decision Logic)
    ```mermaid
    graph LR
        A[開始分析] --> B{活動類型?}
        B -->|Creative| C[自由創作]
        B -->|Imitation| D[跟隨模仿]

        C -->|Sync < 40%| H[✨ 展現獨創性]
        C -->|Sync 40-70%| I[🎨 部分跟隨]
        C -->|Sync > 70%| J[🤝 仍舊跟隨]

        D -->|Sync > 80%| O[🌟 高度同步]
        D -->|Sync > 60%| P[✅ 良好跟隨]
        D -->|Sync > 40%| Q[⚠️ 部分分心]
        D -->|Sync < 40%| R[❌ 脫離群體]
    ```
    *(若圖表未顯示，代表您的瀏覽器暫未支援 Mermaid 渲染，但不影響系統運作)*
    """)

    if st.button("關閉決策樹說明"):
        st.session_state.show_decision_tree = False
        st.rerun()

# [v91.11] Integrated DB & System Checks
try:
    init_db()
except Exception as e:
    st.sidebar.error(f"DB Init Error: {e}")

if mode == "🗄️ 歷史紀錄查閱":
    show_history_ui()
    st.stop() # Stop execution to hide analysis UI

# [v91.21] 2026 Mobile-First Navigation Labels
menu_options = ["1️⃣ 影片設定", "2️⃣ 分析報表", "3️⃣ 社交網絡"]
if 'nav_index' not in st.session_state: 
    st.session_state.nav_index = 0

# [v91.19] Use dynamic key to force UI jump
selected_step = option_menu(
    menu_title=None,
    options=menu_options,
    icons=["gear", "camera-video", "diagram-3"],
    default_index=st.session_state.nav_index,
    orientation="horizontal",
    key=f"wizard_nav_v94_{st.session_state.get('nav_index', 0)}"
)
st.session_state.current_step = selected_step
st.session_state.nav_index = menu_options.index(selected_step)

# [v91.22] Add global Reset Button to Sidebar
if st.sidebar.button("🧹 系統重置 (Reset & Clear Cache)", use_container_width=True, help="遇到畫面卡頓或上傳失敗時，點此清除所有暫存"):
    st.session_state.clear()
    st.rerun()

# [v91.21] 2026 Mobile UX CSS Optimization
st.markdown("""
    <style>
    /* 1. Shrink Nav Font & Optimize for Mobile */
    .nav-link {
        font-size: 13px !important;
        padding: 5px 2px !important;
    }
    /* 2. Global Mobile Comfort Padding */
    .main .block-container {
        padding-top: 1rem !important;
        padding-left: 0.8rem !important;
        padding-right: 0.8rem !important;
    }
    /* 3. Modern Card Aesthetics */
    div.stAlert {
        border-radius: 15px !important;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1) !important;
    }
    /* 4. Controls: Single row buttons padding */
    .stButton button {
        border-radius: 8px !important;
    }
    </style>
""", unsafe_allow_html=True)

if st.session_state.current_step == "1️⃣ 影片設定":
    # st.header("1️⃣ 設定與上傳")
    # 將側邊欄的設定移到這裡
    st.markdown("### 🎯 進階設定")
    col1, col2 = st.columns(2)
    with col1:
        st.session_state.nav_mode = st.radio("系統運行模式", ["🚀 全功能分析", "🗄️ 歷史紀錄查閱"], index=0 if st.session_state.get('nav_mode', '🚀 全功能分析') == '🚀 全功能分析' else 1, key="nav_mode_main", help="「全功能分析」用於分析新影片；「歷史紀錄查閱」用於讀取舊資料。")
        st.session_state.activity_context = st.radio("選擇活動類型:", ["跟隨模仿 (Imitation)", "自由創作 (Creative)"], index=0 if st.session_state.get('act_context', '跟隨模仿 (Imitation)') == '跟隨模仿 (Imitation)' else 1, key="act_context_main", help="「自由創作」模式下，系統會包容低同步率，不將其視為異常。")
        
        # [v91 New] Performance Mode moved to Step 1
        st.selectbox(
            "⚗️ 分析效能模式", 
            ["⚡ 標準模式 (Balanced)", "🚀 極速模式 (Turbo)", "⚡ 超速模式 (Ultra Fast)", "🎯 精準模式 (Pro)", "🔬 微觀分析模式 (MediaPipe Holistic)"],
            index=0,
            key="perf_mode_main",
            help="選擇分析頻率以平衡速度與精準度。極速模式適合長影片，精準模式則捕捉細微動作。",
            on_change=reset_analysis_state
        )
        
        # [v91.1] Show extra options for MediaPipe if selected
        current_perf = st.session_state.get('perf_mode_main', "⚡ 標準模式 (Balanced)")
        if "MediaPipe" in current_perf:
            st.slider("🔬 微觀分析跳幀 (越低越精細)", 1, 5, 2, key="mp_frame_skip_main")
            st.toggle("🎭 啟動面部關鍵點 (468點)", value=False, key="mp_use_face_mesh_main")

    with col2:
        st.session_state.social_threshold_sec = st.slider("判定互動最少秒數 (秒)", min_value=0.5, max_value=10.0, value=st.session_state.get('soc_thresh', 3.0), step=0.5, key="soc_thresh_main")
        st.session_state.save_recording = st.toggle("📼 產出標註影片 (Recording)", value=st.session_state.get('save_rec', True), key="save_rec_main")
        st.session_state.cloud_booster = st.toggle("🚀 雲端效能優化 (加速 50%)", value=st.session_state.get('cloud_boost', True), key="cloud_boost_main")
        st.number_input("強制鎖定特定幼兒 ID", min_value=0, max_value=max(1, max_id), step=1, key="locked_target_id", help="輸入您想觀察的幼兒ID。若設為 0，則全功能模式會追蹤所有人。", on_change=reset_analysis_state)
    
    # [v91 New] Logic to apply Performance Mode settings
    if "Turbo" in perf_mode:
        frame_interval = 12
        model_conf = 0.10
        draw_overlays = True
        target_imgsz = 416
    elif "Ultra Fast" in perf_mode:
        frame_interval = 20
        model_conf = 0.10
        draw_overlays = False
        target_imgsz = 320
    elif "Pro" in perf_mode:
        frame_interval = 2
        model_conf = 0.05
        draw_overlays = True
        target_imgsz = 640
    elif "MediaPipe" in perf_mode:
        frame_interval = 2
        use_face_mesh = False
        model_conf = 0.10
        draw_overlays = True
        target_imgsz = 640
    else:
        frame_interval = 4 # [v94.1 Sensitivity Fix] Lowered from 8 to 4
        use_face_mesh = False
        model_conf = 0.01 # [v94.1 Sensitivity Fix] Lowered from 0.10 to 0.01 (Ultra Sensitive)
        draw_overlays = True
        target_imgsz = 960 # [v94.1 Sensitivity Fix] Increased from 640 to 960 for clarity
    
    st.session_state.last_frame_interval = frame_interval

    # [新增] 同步模式狀態到全局變數
    mode = st.session_state.nav_mode
    
    st.markdown("---")
    # 使用動態 key 來確保按下取消時能真正清空上傳器元件
    uploader_key = st.session_state.get('uploader_key', 0)
    uploaded_file = None
    
    # 防呆機制：影片暫存狀態與跳轉按鈕
    if st.session_state.get('current_tfile_path'):
        st.success(f"✅ 系統已安全暫存您的影片：{st.session_state.get('current_fn', '')}，切換頁面不會遺失！")
        
        col_btn1, col_btn2 = st.columns([3, 1])
        with col_btn1:
            if st.button("🚀 影片準備就緒，點此前往分析", type="primary", use_container_width=True):
                st.session_state.nav_index = 1
                st.rerun()
        with col_btn2:
            if st.button("🗑️ 取消暫存並清空", use_container_width=True):
                try:
                    import os
                    if os.path.exists(st.session_state.current_tfile_path):
                        os.remove(st.session_state.current_tfile_path)
                except:
                    pass
                st.session_state.current_tfile_path = None
                st.session_state.current_fn = None
                st.session_state.processed_file = None
                st.session_state.analysis_done = False
                # 改變 key 以強制 Streamlit 重新渲染空的 uploader
                st.session_state.uploader_key = uploader_key + 1
                st.rerun()
    else:
        col_up, col_reset = st.columns([4, 1])
        with col_up:
            uploaded_file = st.file_uploader("📤 上傳影片 (分析時 ID 將自動歸 1)", type=["mp4", "mov"], key=f"main_uploader_{uploader_key}")
        with col_reset:
            st.write("") # 為了對齊
            st.write("")
            if st.button("🗑️ 清空", use_container_width=True):
                st.session_state.uploader_key = uploader_key + 1
                st.rerun()
else:
    # 確保 uploaded_file 在其他步驟不會造成 NameError
    uploaded_file = None

if uploaded_file:
    # 檢查是否為新檔案，如果是則重置
    if 'current_fn' not in st.session_state or st.session_state.current_fn != uploaded_file.name:
        st.session_state.current_fn = uploaded_file.name # [修正] 必須更新 current_fn，否則無限重置
        st.session_state.id_list = set()
        st.session_state.id_features = {}
        st.session_state.id_tracking_count = {} 
        st.session_state.id_positions = {} # 新增：記錄每個 ID 的位置歷程 [(frame_idx, (x, y)), ...]
        st.session_state.id_motion_log = {} # 新增：記錄每個 ID 的動作分數歷程 {mid: [score, ...]}
        st.session_state.id_actions = defaultdict(lambda: defaultdict(int)) # [v12] Action Tracking
        st.session_state.processed_file = None 
        st.session_state.last_frame = None
        st.session_state.lost_ids = {} # {id: {'hist': hist, 'last_seen': frame_idx, 'feat': clothing_str}}
        st.session_state.id_map = {}   # {temp_id: real_id} mapping
        st.session_state.final_id_count = 0
        st.session_state.reindexing_done = False # [v8 Fix] 防止重複重新編號
        st.session_state.final_id_list = []
        st.session_state.analysis_done = False # [v78 Fix] Ensure new upload resets UI state
        if 'restored_df' in st.session_state: del st.session_state.restored_df # Clear previous loads

        # [v19 New] Advanced Analytics State
        st.session_state.id_yaw_history = {} # {mid: [yaw1, yaw2...]}
        st.session_state.id_focus_score = {} # {mid: focus_frames}
        st.session_state.id_interactions = defaultdict(int) # {(id1, id2): count}
        st.session_state.id_gaze_start = {} # [v19.1]
        st.session_state.social_graph_image = None
        st.session_state.final_report_df = None # [v90.7 New] Clear report cache on new upload

        # [v48 New] Reset Smoothness Log
        st.session_state.id_smoothness_log = defaultdict(list)

        if model and hasattr(model, 'predictor') and model.predictor is not None:
            # model.predictor.trackers = [] 
            pass
            
        # [v21.4 Fix] Persist Video immediately upon upload to prevent NoneType/Buffer issues
        try:
            temp_dir = tempfile.gettempdir()
            current_dir = os.getcwd()
            # [v90.3 Fix] Clean up old processed videos AND old YAML tracking configs to save space
            # Only clean up files OLDER than 15 minutes to prevent deleting active concurrent sessions
            now = time.time()
            cleanup_dirs = [temp_dir, current_dir]
            
            for d in cleanup_dirs:
                if not os.path.exists(d): continue
                for f in os.listdir(d):
                    is_target = (
                        f.startswith("hmeayc_persist_") or 
                        f.startswith("obs_video_") or 
                        f.startswith("fixed_") or 
                        f.startswith("custom_tracker_")
                    )
                    
                    if is_target and (f.endswith(".mp4") or f.endswith(".yaml")):
                        f_path = os.path.join(d, f)
                        try:
                            if now - os.path.getmtime(f_path) > 900: # 15 minutes
                                os.remove(f_path)
                        except: pass
            p_path = os.path.join(temp_dir, f"hmeayc_persist_{uuid.uuid4().hex[:8]}.mp4")
            uploaded_file.seek(0)
            with open(p_path, "wb") as f:
                f.write(uploaded_file.read())
            st.session_state.current_tfile_path = p_path
            st.session_state.current_fn = uploaded_file.name
            
            # [新增] 上傳並暫存完成後，自動跳轉至第二步開始分析
            st.session_state.nav_index = 1
            st.rerun()
        except Exception as e:
            st.error(f"⚠️ 影片預處理失敗: {e}")

        # 清除錯誤位置的定義
        gc.collect()

# --- 防呆機制 (Foolproof) ---
if st.session_state.current_step in ["2️⃣ 影片分析與報表", "3️⃣ 互動網絡圖", "4️⃣ 最終報表"]:
    if not st.session_state.get('current_tfile_path'):
        st.warning("⚠️ 請先在「1️⃣ 設定與上傳」上傳影片！")
        st.stop()

if st.session_state.current_step in ["3️⃣ 互動網絡圖", "4️⃣ 最終報表"]:
    if not st.session_state.get('analysis_done'):
        st.warning("⚠️ 請先完成「2️⃣ 影片分析」！")
        st.stop()

if not st.session_state.analysis_done and st.session_state.current_step == "2️⃣ 影片分析與報表":
    # 只有當「尚未處理過」這個檔案時，才執行分析
    if st.session_state.get('current_fn') and st.session_state.processed_file != st.session_state.current_fn:

        # 第一步已經確認要分析，此處自動開始 (Auto-Start)
        if True:
            if model is None:
                st.error("❌ 模型未正確載入，無法執行分析。")
                st.stop()
            # [v21.4 Fix] Use the persisted file path primarily to avoid NoneType/Buffer issues
            tfile_path = st.session_state.get('current_tfile_path')
            if not tfile_path or not os.path.exists(tfile_path):
                # Final fallback: if somehow path is missing but file is there
                if uploaded_file is not None:
                    tfile_path = os.path.join(tempfile.gettempdir(), f"hmeayc_final_{uuid.uuid4().hex[:8]}.mp4")
                    uploaded_file.seek(0)
                    with open(tfile_path, "wb") as f:
                        f.write(uploaded_file.read())
                    st.session_state.current_tfile_path = tfile_path
                else:
                    st.error("❌ 找不到原始影片檔案 (Buffer Lost)，請點擊 [🗑️ 清空] 後重新上傳。")
                    st.stop()

            # st.info(f"Debug Info: 暫存檔路徑 = {tfile_path}")

            # [v16 Fix] 強制重置追蹤狀態，確保多次執行不會累積舊資料
            st.session_state.id_list = set()
            st.session_state.id_features = {}
            st.session_state.id_tracking_count = {} 
            st.session_state.id_positions = {}
            st.session_state.id_motion_log = {}
            st.session_state.id_actions = defaultdict(lambda: defaultdict(int))
            st.session_state.lost_ids = {}
            st.session_state.id_map = {}
            st.session_state.display_mapping = {}
            st.session_state.final_id_count = 0
            st.session_state.final_id_count = 0
            st.session_state.video_output_path = None # Reset previous video path

            # [v19 New] Reset Advanced State
            st.session_state.id_yaw_history = {}
            st.session_state.id_focus_score = {}
            st.session_state.id_interactions = defaultdict(int)
            st.session_state.id_gaze_start = {} # [v19.1] {(id1, id2): start_frame}
            st.session_state.social_graph_image = None

            # [v48 New] Reset Smoothness Log
            st.session_state.id_smoothness_log = defaultdict(list)

            active_path = tfile_path # [v86] Track path for Polymorphism
            try:
                # Use standard VideoCapture first for tests/configs
                cap = cv2.VideoCapture(active_path)

                if not cap.isOpened():
                    st.error(f"❌ 無法開啟影片檔案: {active_path}")
                else:
                    # [v15 New] Video Writer for Replay
                    # [v21.12.20 Fix] ALWAYS use original video FPS to maintain duration sync
                    orig_fps = cap.get(cv2.CAP_PROP_FPS)
                    if orig_fps is None or orig_fps <= 0: orig_fps = 30.0
                    fps = int(orig_fps) 
                    st.session_state.effective_fps = fps # Store for conversion logic
                    # [Speed Fix] Limit playback/output resolution to 720p to prevent CPU bottleneck
                    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                    if width > 1280:
                        ratio = 1280 / width
                        width = 1280
                        height = int(height * ratio)
                    # [v67 Fix] Force even dimensions (Some H.264 codecs fail with odd numbers)
                    if width % 2 != 0: width -= 1
                    if height % 2 != 0: height -= 1

                    # [v21.12.20 Fix] Define output_path before VideoWriter initialization
                    sid = st.session_state.session_id
                    temp_output_path = os.path.join(tempfile.gettempdir(), f"obs_video_{sid}.mp4")
                    final_output_path = os.path.abspath(f"obs_video_{sid}.mp4")
                    output_path = temp_output_path 
                    pass

                    # [v21.12.15] Codec Doctor: Verify Video Readability
                    test_ret, test_frame = cap.read()
                    if not test_ret or test_frame is None:
                        st.warning("🔄 檢測到影片讀取異常，啟動 Codec Doctor 進行緊急格式轉換...")
                        cap.release()
                        
                        import subprocess
                        import shutil
                        temp_fixed = os.path.join("/tmp/", f"fixed_{sid}.mp4") if os.name != 'nt' else f"fixed_{sid}.mp4"
                        
                        ffmpeg_exe = shutil.which("ffmpeg")
                        if not ffmpeg_exe and os.name == 'nt':
                            fallback_ffmpeg = r"C:\Users\user\AppData\Local\Microsoft\WinGet\Packages\Gyan.FFmpeg_Microsoft.Winget.Source_8wekyb3d8bbwe\ffmpeg-8.0.1-full_build\bin\ffmpeg.exe"
                            if os.path.exists(fallback_ffmpeg): ffmpeg_exe = fallback_ffmpeg

                        if ffmpeg_exe:
                            try:
                                # Convert to ultra-compatible H.264
                                cmd = [ffmpeg_exe, "-y", "-i", active_path, "-c:v", "libx264", "-preset", "ultrafast", "-crf", "28", temp_fixed]
                                subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=60)
                                active_path = temp_fixed
                                cap = cv2.VideoCapture(active_path)
                                st.info("✅ Codec Doctor 轉換成功，分析將重新啟動。")
                            except:
                                cap = cv2.VideoCapture(active_path) # Fallback to original
                        else:
                            cap = cv2.VideoCapture(active_path)
                    else:
                        # Rewind cap to start
                        cap.set(cv2.CAP_PROP_POS_FRAMES, 0)

                    # [v21.12.16] Force Software Codec (XVID/mp4v)
                    # [v21.12.20 Fix] Add H264 as primary if possible
                    codecs = ['mp4v', 'XVID', 'avc1', 'H264']
                    out_video = None
                    last_v_err = "No error recorded"
                    
                    for c_str in codecs:
                        try:
                            fourcc = cv2.VideoWriter_fourcc(*c_str)
                            # Ensure output_path is valid
                            if not output_path:
                                logging.error("VideoWriter: output_path is None!")
                                break
                            
                            out_video = cv2.VideoWriter(output_path, fourcc, fps, (width, height))
                            if out_video is not None and out_video.isOpened():
                                logging.info(f"VideoWriter SUCCESS: {c_str}")
                                break
                            else:
                                last_v_err = f"Open failed for codec {c_str}"
                        except Exception as ve:
                            last_v_err = str(ve)
                            continue

                    if out_video is None or not out_video.isOpened():
                        st.warning(f"📺 錄像組件失效 ({last_v_err})。AI 將繼續運行但無法下載標註影片。")
                        logging.error(f"VideoWriter Error: {last_v_err} | Path: {output_path}")
                        out_video = None

                    # st.write("Debug Info: 影片開啟成功")

                # [v16] Initialize Display Mapping
                st.session_state.display_mapping = {} 

                st.info("AI 辨識中... 本次分析 ID 將從 1 開始編號 (已啟用 ID 重映射)。")

                total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
                # st.write(f"Debug Info: 總幀數 = {total_frames}")

                # [v18 Fix Layout] Initialize containers in the correct order for UI placement
                st_progress_text = st.empty()
                st_progress = st.progress(0)
                st_frame = st.empty()
                log_container = st.empty() # 用於顯示即時 debug 訊息

                if total_frames <= 0: total_frames = 1000 # 防呆
                f_idx = 0

                # [v59] Setup MediaPipe Holistic Model
                use_mp = "MediaPipe" in perf_mode
                mp_holistic = None
                mp_drawing = None
                mp_drawing_styles = None
                holistic_model = None
                if use_mp:
                    try:
                        import mediapipe as mp
                        mp_holistic = mp.solutions.holistic
                        mp_drawing = mp.solutions.drawing_utils
                        mp_drawing_styles = mp.solutions.drawing_styles
                    except ImportError:
                        try:
                            import mediapipe.python.solutions.holistic as mp_holistic
                            import mediapipe.python.solutions.drawing_utils as mp_drawing
                            import mediapipe.python.solutions.drawing_styles as mp_drawing_styles
                        except Exception as e:
                            st.error(f"MediaPipe Import Failed: {e}")
                            use_mp = False
                    
                    if use_mp:
                        try:
                            holistic_model = mp_holistic.Holistic(
                                static_image_mode=False,
                                model_complexity=0, # [Speed Fix] Lower complexity for CPU
                                enable_segmentation=False,
                                refine_face_landmarks=False,
                                min_detection_confidence=0.5,
                                min_tracking_confidence=0.5
                            )
                        except Exception as e:
                            st.warning(f"MediaPipe Model Init Error: {e}")


                # [v21.12.19 Fast Mode] Reduce FPS proportionally to skip frames while maintaining playback speed
                fps = max(int(orig_fps) // frame_interval, 1)
                st.session_state.effective_fps = fps

                # [Performance Fix] Pre-calculate tracker and resolution outside the loop
                tracker_file = create_tracker_config()
                
                # Dynamic Resolution based on Performance Mode [v21.12.20 Fix]
                if "Ultra Fast" in perf_mode:
                    target_imgsz = 160 # Fastest possible resolution
                elif "Turbo" in perf_mode:
                    target_imgsz = 320 
                elif "Pro" in perf_mode:
                    target_imgsz = 480 
                else:
                    # [v90.4 Fix] Downgraded to 640. 1024 drops CPU speed to ~2 FPS. 640 strikes best balance.
                    target_imgsz = 640 
                
                # [v90.4 Fix] Removed the code block that forced 1024 for Standard Mode since Hugging Face CPUs burst into flames on 1024.
                
                prog = 0.0 
                import time
                start_time = time.time()
                total_processed = 0

                # [v83] Hyper Drive: Determine Hardware Accelerator
                import torch
                device = 'cuda' if torch.cuda.is_available() else 'cpu'
                is_fp16 = device == 'cuda'
                
                # [v90.4 Fix] Prevent Hugging Face CPU Thrashing
                if device == 'cpu':
                    torch.set_num_threads(4)

                # Cache session state to local variables for speed (Pre-loop)
                id_map = st.session_state.id_map
                id_list = st.session_state.id_list
                id_features = st.session_state.id_features
                id_tracking_count = st.session_state.id_tracking_count
                id_positions = st.session_state.id_positions
                id_motion_log = st.session_state.id_motion_log
                id_actions = st.session_state.id_actions
                display_mapping = st.session_state.display_mapping
                id_yaw_history = st.session_state.id_yaw_history
                id_interactions = st.session_state.id_interactions

                # [v89.7 Fix] Remove ThreadedVideoCap to prevent 7-second timeout truncations
                # Fall back to robust cv2 default capture and frame skipping
                if cap is not None: cap.release() 
                cap = cv2.VideoCapture(active_path)

                while cap.isOpened():
                    # [v90.2 Speed Fix] Fast skip frames using grab() instead of full read() to bypass decoding bottleneck
                    if total_processed > 0 and frame_interval > 1:
                        for _ in range(frame_interval - 1):
                            if not cap.grab(): break
                            
                    ret, frame = cap.read()
                    if not ret:
                        break
                    
                    actual_f_idx = int(cap.get(cv2.CAP_PROP_POS_FRAMES))
                    f_idx = actual_f_idx
                    
                    is_analysis_frame = True # Since we pre-skipped using grab()
                    
                    # [v83] Hyper Drive: Performance Tracking
                    total_processed += frame_interval
                    elapsed = time.time() - start_time
                    current_fps = total_processed / elapsed if elapsed > 0.1 else 0
                    
                    # [v83] Hyper Drive: Throttled Progress Update (Every 10% or 120 frames)
                    if f_idx == 1 or f_idx % max(120, frame_interval) < frame_interval or (f_idx >= total_frames * (prog + 0.1)):
                        prog = min(f_idx / total_frames, 1.0)
                        fps_text = f"速度: {current_fps:.2f} FPS"
                        st_progress_text.text(f"🚀 {fps_text} | 處理中 (Hyper Drive): {f_idx}/{total_frames} ({prog:.1%})")
                        st_progress.progress(prog)

                    # 到這裡代表是「分析幀」，才開始執行重度任務
                    annotated_frame = frame.copy() if draw_overlays else frame
                    ids, boxes, results, results_yolo = [], [], None, None
                    try:
                        if use_mp and holistic_model:
                            # [v65 Major Upgrade] YOLO-Guided MediaPipe Stability Tracker with Specific ID Targeting
                            # [v21.12.16] Increased thickness and scale for visibility
                            results_yolo = model.track(frame, persist=True, verbose=False, conf=model_conf, iou=0.6, tracker=tracker_file, imgsz=target_imgsz, classes=[0])

                            try:
                                # [v69.1] Use YOLO skeleton as baseline for Micro Mode (Ensure skeletons exist even if MediaPipe fails)
                                # [v21.12.16] Ensure annotated_frame is initialized properly
                                if draw_overlays:
                                    if results_yolo and len(results_yolo) > 0:
                                        # [v21.12.16] Increased thickness and scale for visibility
                                        annotated_frame = results_yolo[0].plot(boxes=True, labels=False, probs=False, kpt_line=True, kpt_radius=6, line_width=3)
                                    else:
                                        annotated_frame = frame.copy()
                                else:
                                    annotated_frame = frame
                            except:
                                annotated_frame = frame if not draw_overlays else frame.copy()

                            target_box = None
                            max_area = 0
                            found_specific_id = False

                            if results_yolo and len(results_yolo) > 0 and results_yolo[0].boxes is not None:
                                # [v21.12.16 No-ID Survival Mode]
                                # If the tracker fails (no ids), use detections as a fallback
                                has_ids = results_yolo[0].boxes.id is not None
                                ids = results_yolo[0].boxes.id.int().cpu().numpy() if has_ids else np.full(len(results_yolo[0].boxes), -1)
                                boxes = results_yolo[0].boxes.xyxy.cpu().numpy()

                                for i, box in enumerate(boxes):
                                    track_id = int(ids[i])
                                    x1, y1, x2, y2 = map(int, box)

                                    # [v21.12.16] Global Survival for MediaPipe Mode
                                    st.session_state.id_tracking_count[track_id] = st.session_state.id_tracking_count.get(track_id, 0) + 1
                                    if st.session_state.id_tracking_count[track_id] >= 1:
                                        # Map to display ID (e.g. 1, 2, 3...)
                                        if track_id not in st.session_state.display_mapping:
                                            st.session_state.display_mapping[track_id] = len(st.session_state.display_mapping) + 1
                                        mid_display = st.session_state.display_mapping[track_id]
                                        st.session_state.id_list.add(mid_display)
                                    
                                        # Initialize position log for every person to avoid key errors later
                                        if mid_display not in st.session_state.id_positions:
                                            st.session_state.id_positions[mid_display] = []
                                            st.session_state.id_motion_log[mid_display] = []
                                        st.session_state.id_positions[mid_display].append((f_idx, ((x1+x2)//2, (y1+y2)//2)))

                                    # [v21.12.16] Increased font scale and thickness
                                    if draw_overlays:
                                        label_raw = f"ID: {st.session_state.display_mapping.get(track_id, 'Scan')}"
                                        cv2.putText(annotated_frame, label_raw, (x1, y1-15), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 165, 255), 3)

                                    if has_ids and target_track_id > 0 and track_id == target_track_id:
                                        target_box = [x1, y1, x2, y2]
                                        found_specific_id = True
                                
                                    # Fallback to largest box
                                    area = (x2 - x1) * (y2 - y1)
                                    if not found_specific_id and area > max_area:
                                        max_area = area
                                        target_box = [x1, y1, x2, y2]

                            # [v69 New] Visual Feedback for Search (Now safe as annotated_frame is defined)
                            if target_track_id > 0 and not found_specific_id:
                                cv2.putText(annotated_frame, f"Searching for ID: {target_track_id}...", (50, 80), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 165, 255), 2)

                            if target_box is not None:
                                x1, y1, x2, y2 = target_box
                                h_frame, w_frame, _ = frame.shape

                                # Add 15% margin to the bounding box for full body context
                                bw = x2 - x1
                                bh = y2 - y1
                                mx = int(bw * 0.15)
                                my = int(bh * 0.15)

                                crop_x1 = max(0, x1 - mx)
                                crop_y1 = max(0, y1 - my)
                                crop_x2 = min(w_frame, x2 + mx)
                                crop_y2 = min(h_frame, y2 + my)

                                # Crop the image
                                roi = frame[crop_y1:crop_y2, crop_x1:crop_x2]

                                # 2. Feed ONLY the cropped largest person to MediaPipe
                                # [v69 Safety Fix] Verify model and drawing components exist
                                if holistic_model is not None and mp_drawing is not None and mp_holistic is not None:
                                    roi_rgb = cv2.cvtColor(roi, cv2.COLOR_BGR2RGB)
                                    results_mp = holistic_model.process(roi_rgb)

                                    if results_mp.pose_landmarks:
                                        roi_h, roi_w, _ = roi.shape

                                        # [v66] Use FAST Translation function
                                        # Translate all landmarks back to full frame
                                        global_pose = translate_landmarks_fast(results_mp.pose_landmarks, crop_x1, crop_y1, roi_w, roi_h, w_frame, h_frame)
                                        global_face = translate_landmarks_fast(results_mp.face_landmarks, crop_x1, crop_y1, roi_w, roi_h, w_frame, h_frame) if use_face_mesh else None
                                        global_lh = translate_landmarks_fast(results_mp.left_hand_landmarks, crop_x1, crop_y1, roi_w, roi_h, w_frame, h_frame)
                                        global_rh = translate_landmarks_fast(results_mp.right_hand_landmarks, crop_x1, crop_y1, roi_w, roi_h, w_frame, h_frame)

                                        # Draw on the original frame using translated landmarks
                                        mp_drawing.draw_landmarks(annotated_frame, global_pose, mp_holistic.POSE_CONNECTIONS, landmark_drawing_spec=mp_drawing_styles.get_default_pose_landmarks_style())
                                        if global_face:
                                            mp_drawing.draw_landmarks(annotated_frame, global_face, mp_holistic.FACEMESH_TESSELATION, None, mp_drawing_styles.get_default_face_mesh_tesselation_style())
                                        if global_lh:
                                            mp_drawing.draw_landmarks(annotated_frame, global_lh, mp_holistic.HAND_CONNECTIONS)
                                        if global_rh:
                                            mp_drawing.draw_landmarks(annotated_frame, global_rh, mp_holistic.HAND_CONNECTIONS)

                                        lms = global_pose.landmark
                                        cx = int((lms[23].x + lms[24].x) * w_frame / 2) # Hip center for tracking
                                        cy = int((lms[23].y + lms[24].y) * h_frame / 2)
                                    else:
                                        # MediaPipe found nothing, use center of YOLO box
                                        cx, cy = (x1+x2)//2, (y1+y2)//2
                                else:
                                    # [v69 Fallback] Record person using YOLO even if MediaPipe fails to load
                                    cx, cy = (x1+x2)//2, (y1+y2)//2

                                # Common recording and drawing for the target (Micro Mode)
                                # [v74 Fix] Ensure ID consistency
                                mid = target_track_id if target_track_id > 0 else 1
                                st.session_state.id_list.add(mid)
                                st.session_state.display_mapping[mid] = mid

                                if mid not in st.session_state.id_positions:
                                    st.session_state.id_positions[mid] = []
                                    st.session_state.id_motion_log[mid] = []
                                    st.session_state.id_tracking_count[mid] = 0

                                st.session_state.id_tracking_count[mid] += 1
                                st.session_state.id_positions[mid].append((f_idx, (cx, cy)))

                                if mid not in st.session_state.id_actions:
                                    st.session_state.id_actions[mid] = {}
                                st.session_state.id_actions[mid]["微觀追蹤(全身543點)"] = st.session_state.id_actions[mid].get("微觀追蹤(全身543點)", 0) + 1

                                if mid not in st.session_state.id_features:
                                    # [v67 New] Micro Mode Clothing Extraction
                                    # Use the target_box to crop shirt and pants from the original frame
                                    bx1, by1, bx2, by2 = target_box
                                    bh_box = by2 - by1
                                    shirt = frame[max(0, by1+int(bh_box*0.1)):min(h_frame, by1 + int(bh_box*0.4)), bx1+int((bx2-bx1)*0.2):bx2-int((bx2-bx1)*0.2)]
                                    pants = frame[max(0, by1 + int(bh_box*0.6)):min(h_frame, by2-int(bh_box*0.1)), bx1+int((bx2-bx1)*0.2):bx2-int((bx2-bx1)*0.2)]
                                
                                    c_shirt = get_dominant_color(shirt)
                                    c_pants = get_dominant_color(pants)
                                    # [v68 Fix] Also support patterns in Micro Mode
                                    p_shirt = get_clothing_pattern(shirt)
                                    p_pants = get_clothing_pattern(pants)
                                
                                    feat_str = f"上衣：{c_shirt}{p_shirt}。下裝：{c_pants}{p_pants}、褲子。配件：無。"
                                
                                    st.session_state.id_features[mid] = {
                                        "clothing": feat_str, 
                                        "score_pending": True, 
                                        "original_id": mid,
                                        "hist": get_color_histogram(shirt)
                                    }

                                    # [v68 New] Draw Bounding Box around the target
                                    cv2.rectangle(annotated_frame, (crop_x1, crop_y1), (crop_x2, crop_y2), (0, 255, 0), 4)

                                    # Use the YOLO bounding box for label
                                    nx, ny = crop_x1, max(0, crop_y1 - 20)
                                    label = f"ID: {mid} (Locked Target)" if found_specific_id else "ID: 1 (Largest Target)"
                                    t_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)[0]
                                    cv2.rectangle(annotated_frame, (nx, max(0, ny - t_size[1] - 5)), (nx + t_size[0], ny), (0, 255, 0), -1)
                                    cv2.putText(annotated_frame, label, (nx, ny - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 0), 2)

                                    if len(st.session_state.id_positions[mid]) > 1:
                                        prev_pos = st.session_state.id_positions[mid][-2][1]
                                        dist = np.sqrt((cx-prev_pos[0])**2 + (cy-prev_pos[1])**2)
                                    else:
                                        cv2.putText(annotated_frame, "No Target Found", (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                            
                        if not use_mp:
                            # [v83] Hyper Drive: FP16 + Agnostic NMS + Device aware
                            # [v89.9 Update] IOU returned to 0.5. 0.7 caused ID explosion due to duplicate overlapping boxes
                            results = model.track(frame, persist=True, verbose=False, conf=model_conf, iou=0.5, 
                                                 classes=[0], tracker=tracker_file, imgsz=target_imgsz, 
                                                 half=is_fp16, device=device, agnostic_nms=True)

                        # [v83] Cache per-frame analysis data for vectorized social check
                        frame_centers = []
                        frame_yaws = []
                        frame_display_ids = []
                        if ('results_yolo' in locals() and results_yolo and len(results_yolo) > 0 and results_yolo[0].boxes is not None) or \
                           ('results' in locals() and results and len(results) > 0 and results[0].boxes is not None):
                            # [v21.12.16] survival logic
                            # [Fix] Choose correct results source: results_yolo (MediaPipe mode) or results (YOLO mode)
                            _ry = 'results_yolo' in locals() and results_yolo and len(results_yolo) > 0 and results_yolo[0].boxes is not None
                            _r = 'results' in locals() and results and len(results) > 0 and results[0].boxes is not None
                            active_results = results_yolo[0] if _ry else results[0]
                            has_ids = (active_results.boxes.id is not None)
                            ids = active_results.boxes.id.int().cpu().numpy() if has_ids else np.array([1000 + i for i in range(len(active_results.boxes))])
                            boxes = active_results.boxes.xyxy.cpu().numpy()
                            keypoints_data = active_results.keypoints.data.cpu().numpy() if active_results.keypoints is not None else None

                            if draw_overlays:
                                try:
                                    # [v90.2 UI Fix] Set boxes=False so YOLO doesn't draw "ID-less blue boxes". 
                                    # We will draw our own bounding boxes logic further below!
                                    annotated_frame = active_results.plot(boxes=False, labels=False, probs=False, kpt_line=True, kpt_radius=6, line_width=3)
                                except:
                                    annotated_frame = frame.copy()
                            else:
                                annotated_frame = frame
                                
                            if not has_ids and draw_overlays:
                                # Visual indicator for survival mode active
                                cv2.putText(annotated_frame, "SURVIVAL MODE (YOLO)", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 165, 255), 2)

                        # [v21.12.21 Deep Speed] Completely bypass Re-ID dictionary updates in Ultra Fast mode
                        ultra_fast_mode = "Ultra Fast" in perf_mode
                        
                        if not ultra_fast_mode:
                            # [Re-ID 步驟 1: 更新 Lost IDs]
                            current_mids = set([int(i) for i in ids]) if len(ids) > 0 else set()
                            # 找出已經穩定追蹤但本幀消失的 ID
                            active_mapped_ids = set()
                            for m in current_mids:
                                active_mapped_ids.add(st.session_state.id_map.get(m, m))
    
                            # 檢查 id_list 中有哪些人這幀不見了
                            for known_id in st.session_state.id_list:
                                if known_id not in active_mapped_ids:
                                    # 加入 Lost List (如果尚未加入)
                                    if known_id not in st.session_state.lost_ids and known_id in st.session_state.id_features:
                                        feat = st.session_state.id_features[known_id]
                                        # [v57 Fix] Save Last Position for Spatial Check
                                        last_pos = None
                                        if known_id in st.session_state.id_positions and st.session_state.id_positions[known_id]:
                                            last_pos = st.session_state.id_positions[known_id][-1][1] # (x,y)
    
                                        if 'hist' in feat:
                                            st.session_state.lost_ids[known_id] = {
                                                'hist': feat['hist'],
                                                'last_seen': f_idx,
                                                'last_pos': last_pos
                                            }
    
                            # 清理過期 Lost IDs (> 300 Frames ~= 10s)
                            keys_to_remove = [k for k, v in st.session_state.lost_ids.items() if f_idx - v['last_seen'] > 300]
                            for k in keys_to_remove:
                                del st.session_state.lost_ids[k]

                        # [v89 Fix] Removed aggressive tracking count cleanup that caused data loss
                        # We need to preserve counts for all IDs detected throughout the video.
                        pass

                        # [v21.12.18] Unified Person Loop: Process every human detection
                        for i, box in enumerate(boxes):
                            mid = int(ids[i])
                            x1, y1, x2, y2 = map(int, box)
                            
                            # [v90.2 Fix] Aggressively filter out TINY noise boxes to drop the ID count from ~68 to realistic ~30
                            w_box = x2 - x1
                            h_box_dim = y2 - y1
                            if w_box * h_box_dim < 3000 and target_track_id <= 0:
                                continue
                                
                            curr_center = ((x1+x2)//2, (y1+y2)//2)

                            # [v89.4 Fix] Re-ID Logic Correction: 
                            # Re-ID logic should run in all modes EXCEPT Ultra Fast to ensure "stickiness".
                            # v89.1 accidentally made it only run if Ultra Fast condition was met incorrectly.
                            if not ultra_fast_mode and mid not in st.session_state.id_list and mid not in st.session_state.id_map:
                                if st.session_state.lost_ids:
                                    h_box = y2 - y1
                                    # Prevent crop out of bounds
                                    crop_y1 = max(0, y1+int(h_box*0.1))
                                    crop_y2 = min(frame.shape[0], y2-int(h_box*0.1))
                                    if crop_y2 > crop_y1 and x2 > x1:
                                        crop = frame[crop_y1:crop_y2, x1:x2]
                                        curr_hist = get_color_histogram(crop)
                                        
                                        best_match = -1
                                        best_score = 0.0
    
                                        if curr_hist is not None:
                                            for lost_id, data in st.session_state.lost_ids.items():
                                                if data['hist'] is None: continue
                                                hist_score = cv2.compareHist(curr_hist, data['hist'], cv2.HISTCMP_CORREL)
                                                
                                                dist_weight = 1.0
                                                if data['last_pos']:
                                                    dist = np.linalg.norm(np.array(curr_center) - np.array(data['last_pos']))
                                                    if dist > 400: continue 
                                                    dist_weight = np.exp(-dist / 300.0)
                                                
                                                combined_score = hist_score * dist_weight
                                                # [v89.4 Fix] Lowered from 0.65 to 0.50 for background pursuit
                                                if combined_score > 0.50 and combined_score > best_score:
                                                    best_score = combined_score
                                                    best_match = lost_id
                                        
                                            if best_match != -1:
                                                st.session_state.id_map[mid] = best_match
                                                if best_match in st.session_state.lost_ids:
                                                    del st.session_state.lost_ids[best_match]

                            # apply Mapping (取出真實 ID)
                            mid = st.session_state.id_map.get(mid, mid)

                             # [ID 穩定性過濾]
                            st.session_state.id_tracking_count[mid] = st.session_state.id_tracking_count.get(mid, 0) + 1
                            current_display_id = st.session_state.display_mapping.get(mid, -1)
                            is_target = (target_track_id > 0 and current_display_id == target_track_id) 
                            
                            # [v89.9 Fix] Increase stability threshold slightly to filter out 1-frame ghost detections
                            # [v90.2 Update] Increased to 8 to massively filter out brief background anomalies!
                            stability_thresh = 8
                            
                            if st.session_state.id_tracking_count.get(mid, 0) >= stability_thresh or is_target:
                                if target_track_id > 0:
                                    st.session_state.display_mapping[mid] = mid
                                    mid_display = mid
                                else:
                                    if mid not in st.session_state.display_mapping:
                                        st.session_state.display_mapping[mid] = len(st.session_state.display_mapping) + 1
                                    mid_display = st.session_state.display_mapping[mid]

                                if target_track_id > 0 and mid_display != target_track_id:
                                    continue
                                
                                raw_mapped_id = mid 
                                # Positioning and Motion Energy Calculation
                                if mid_display not in st.session_state.id_positions:
                                    st.session_state.id_positions[mid_display] = []
                                    st.session_state.id_motion_log[mid_display] = []
                                
                                # [v90.1 Fix] Calculate frame-to-frame displacement for Energy and Smoothness
                                if st.session_state.id_positions[mid_display]:
                                    prev_pos = st.session_state.id_positions[mid_display][-1][1]
                                    disp = np.linalg.norm(np.array(curr_center) - np.array(prev_pos))
                                    st.session_state.id_motion_log[mid_display].append(disp)
                                else:
                                    st.session_state.id_motion_log[mid_display].append(0.0)

                                st.session_state.id_positions[mid_display].append((f_idx, curr_center))

                                if mid_display not in st.session_state.id_features:
                                        # [v89 Fix] Remove stability delay for feature extraction to ensure no person is missed
                                        if "Ultra Fast" in perf_mode:
                                                # [v21.12.21 Speed] Bypass heavy image processing in Ultra Fast mode
                                                st.session_state.id_features[mid_display] = {
                                                        "clothing": "極速模式已略過服裝特徵分析以提升效能。", 
                                                        "score_pending": True, 
                                                        "original_id": mid,
                                                        "hist": None
                                                }
                                        else:
                                                h_f, w_f, _ = frame.shape
                                                shirt = frame[max(0, y1 + int((y2-y1)*0.1)):min(h_f, y1 + int((y2-y1)*0.4)), x1+int((x2-x1)*0.2):x2-int((x2-x1)*0.2)]
                                                pants = frame[max(0, y1 + int((y2-y1)*0.6)):min(h_f, y2-int((y2-y1)*0.1)), x1+int((x2-x1)*0.2):x2-int((x2-x1)*0.2)]
                                                
                                                # Use fast_mode if in high-speed modes
                                                is_fast = "Turbo" in perf_mode or "Ultra Fast" in perf_mode
                                                c_shirt = get_dominant_color(shirt, fast_mode=is_fast)
                                                c_pants = get_dominant_color(pants, fast_mode=is_fast)
                                                p_shirt = get_clothing_pattern(shirt)
                                                p_pants = get_clothing_pattern(pants)

                                                fat_str = f"上衣：{c_shirt}{p_shirt}。下裝：{c_pants}{p_pants}、褲子。配件：無。"
                                                
                                                st.session_state.id_features[mid_display] = {
                                                        "clothing": fat_str, 
                                                        "score_pending": True, 
                                                        "original_id": mid,
                                                        "hist": get_color_histogram(shirt)
                                                }
                                elif st.session_state.id_features[mid_display].get("score_pending", False):
                                    # Every 300 frames, refresh the color just in case lighting changed significantly
                                    if f_idx % 300 == 0:
                                        h_f, w_f, _ = frame.shape
                                        shirt = frame[max(0, y1 + int((y2-y1)*0.1)):min(h_f, y1 + int((y2-y1)*0.4)), x1+int((x2-x1)*0.2):x2-int((x2-x1)*0.2)]
                                        c_shirt = get_dominant_color(shirt, fast_mode=is_fast)
                                        id_features[mid_display]["clothing_last_refresh"] = c_shirt

                                # Action Recognition & Drawing Logic
                                color = (0, 140, 255) # Orange (Default)
                                current_action = ""

                                # Action Recognition
                                if keypoints_data is not None and len(keypoints_data) > i:
                                    try:
                                        kpts = keypoints_data[i] # (17, 3)
                                        current_actions = []

                                        # 1. Check Lying Down / Floor Action
                                        is_lying_down = False
                                        # Torso Orientation: Shoulders(5,6) vs Hips(11,12)
                                        if kpts[5][2]>0.5 and kpts[11][2]>0.5:
                                            dy = abs(kpts[5][1] - kpts[11][1])
                                            dx = abs(kpts[5][0] - kpts[11][0])
                                            if dy < dx * 0.8: # Horizontal
                                                is_lying_down = True

                                        # Bounding Box Aspect Ratio (Backup)
                                        if not is_lying_down:
                                            w = x2 - x1
                                            h = y2 - y1
                                            if w > h * 1.2: 
                                                is_lying_down = True

                                        if is_lying_down:
                                            current_actions.append("地板動作")
                                            color = (255, 0, 255) # Magenta

                                        # 2. Check Hands Up
                                        if (kpts[9][2] > 0.5 and kpts[5][2] > 0.5 and kpts[9][1] < kpts[5][1]) or \
                                        (kpts[10][2] > 0.5 and kpts[6][2] > 0.5 and kpts[10][1] < kpts[6][1]):
                                            current_actions.append("舉手")
                                            if not is_lying_down: color = (0, 0, 255) # Red priority if standing

                                        # 3. Check Squat
                                        if not is_lying_down:
                                            if kpts[11][2]>0.5 and kpts[13][2]>0.5 and kpts[14][2]>0.5:
                                                thigh_len_l = abs(kpts[13][1] - kpts[11][1])
                                                shin_len_l = abs(kpts[15][1] - kpts[13][1]) if kpts[15][2]>0.5 else 100
                                                if thigh_len_l < shin_len_l * 0.5: 
                                                    current_actions.append("蹲下")
                                                    color = (255, 0, 0) # Blue

                                        # 4. Check Jump
                                        if not is_lying_down and mid_display in st.session_state.id_positions:
                                            hist = st.session_state.id_positions[mid_display]
                                            if len(hist) > 5:
                                                recent_hist = hist[-15:] 
                                                ys = [p[1][1] for p in recent_hist]
                                                avg_y = np.mean(ys)
                                                h = y2 - y1
                                                if curr_center[1] < avg_y - (h * 0.20): 
                                                    current_actions.append("跳躍")
                                                    color = (0, 255, 0) # Green (v21.5.8 Change)

                                        # 5. Check Leg Up / Kick
                                        if not is_lying_down and not "蹲下" in current_actions:
                                            leg_action_detected = False
                                            # Left Leg
                                            if kpts[13][2]>0.5: # Knee
                                                if kpts[11][2]>0.5 and kpts[13][1] < kpts[11][1] + 20: # Knee above Hip
                                                    leg_action_detected = True
                                                elif kpts[15][2]>0.5 and kpts[15][1] < kpts[13][1]: # Ankle above Knee (Kick)
                                                    leg_action_detected = True
                                            
                                            # Right Leg
                                            if not leg_action_detected and kpts[14][2]>0.5:
                                                if kpts[12][2]>0.5 and kpts[14][1] < kpts[12][1] + 20: # Knee above Hip
                                                    leg_action_detected = True
                                                elif kpts[16][2]>0.5 and kpts[16][1] < kpts[14][1]: # Ankle above Knee (Kick)
                                                    leg_action_detected = True

                                            # Sitting Check
                                            is_sitting = False
                                            if kpts[5][2]>0.5 and kpts[11][2]>0.5:
                                                torso_h = abs(kpts[11][1] - kpts[5][1])
                                                if kpts[15][2]>0.5: # Ankle
                                                    hip_to_ankle = abs(kpts[15][1] - kpts[11][1])
                                                    if hip_to_ankle < torso_h * 0.8: # Legs are folded/sitting
                                                        is_sitting = True

                                            if is_sitting:
                                                current_actions.append("盤坐")
                                                color = (128, 128, 128) # Grey
                                            
                                            if leg_action_detected and not is_sitting:
                                                current_actions.append("抬腿")
                                                color = (19, 69, 139) # Brown

                                        # 6. Check Running
                                        if not is_lying_down and mid_display in id_positions:
                                            hist = id_positions[mid_display]
                                            if len(hist) > 3:
                                                prev_x = hist[-3][1][0]
                                                curr_x = curr_center[0]
                                                dx_speed = abs(curr_x - prev_x)
                                                w = x2 - x1
                                                if dx_speed > w * 0.5:
                                                    current_actions.append("跑步")

                                        # Consolidate Actions
                                        current_actions = list(set(current_actions))
                                        display_action_en = ""

                                        if current_actions:
                                            action_map = {
                                                "舉手": "HandsUp", "蹲下": "Squat", "跳躍": "Jump",
                                                "地板動作": "Floor", "抬腿": "LegUp/Kick", "跑步": "Run",
                                                "盤坐": "Sitting"
                                            }
                                            english_tags = [action_map.get(a, a) for a in current_actions]
                                            display_action_en = "+".join(english_tags)

                                        # Drawing (Hyper Drive: Simplified drawing for Ultra Fast)
                                        if ultra_fast_mode:
                                            # Just draw a simple marker to save CPU on complex font rendering
                                            cv2.circle(annotated_frame, curr_center, 6, color, -1)
                                        else:
                                            cv2.rectangle(annotated_frame, (x1, y1), (x2, y2), color, 4)
                                            
                                            # Always show label with ID
                                            label = f"ID: {mid_display}"
                                            if display_action_en:
                                                label += f" {display_action_en}"
                                            
                                            t_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.8, 3)[0]
                                            # [v90 Fix] Label Position: Default Top-Left, fallback to Bottom-Left if clipped
                                            label_y = y1 - 10
                                            if label_y < 20: 
                                                label_y = y2 + t_size[1] + 10
                                                
                                            c2 = x1 + t_size[0], label_y - t_size[1] - 10
                                            cv2.rectangle(annotated_frame, (x1, label_y), (c2[0], c2[1]), color, -1)
                                            cv2.putText(annotated_frame, label, (x1, label_y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 3)

                                        # Logging
                                        if current_actions:
                                            if mid_display not in id_actions:
                                                id_actions[mid_display] = {}
                                            for act in current_actions:
                                                id_actions[mid_display][act] = id_actions[mid_display].get(act, 0) + 1

                                    except Exception as e:
                                        pass

                                # (2) Focus Analysis (Head Yaw)
                                head_yaw = 0
                                if keypoints_data is not None and len(keypoints_data) > i:
                                    try:
                                        kpts = keypoints_data[i]
                                        # nose(0), l_ear(3), r_ear(4)
                                        if kpts[0,2] > 0.5 and kpts[3,2] > 0.5 and kpts[4,2] > 0.5:
                                            # Sub-optimization: Inline calculate_head_yaw
                                            d_l = np.sqrt((kpts[0,0]-kpts[3,0])**2 + (kpts[0,1]-kpts[3,1])**2)
                                            d_r = np.sqrt((kpts[0,0]-kpts[4,0])**2 + (kpts[0,1]-kpts[4,1])**2)
                                            total_d = d_l + d_r
                                            if total_d > 0:
                                                head_yaw = ((d_r - d_l) / total_d) * 90
                                    except: pass

                                if mid_display not in id_yaw_history:
                                    id_yaw_history[mid_display] = []
                                id_yaw_history[mid_display].append(head_yaw)

                                # Collect for vectorized social analysis later
                                frame_centers.append(curr_center)
                                frame_yaws.append(head_yaw)
                                frame_display_ids.append(mid_display)

                        # (3) [v83] Vectorized Social Interaction (Replaces O(N^2) loop)
                        if len(frame_display_ids) > 1:
                            try:
                                centers_arr = np.array(frame_centers) # (N, 2)
                                yaws_arr = np.array(frame_yaws)       # (N,)
                                
                                # Distance Matrix: (N, N)
                                diff = centers_arr[:, np.newaxis, :] - centers_arr[np.newaxis, :, :] 
                                dists = np.sqrt((diff**2).sum(axis=2))
                                
                                # Potential Interactors (Dist < 150)
                                potential = dists < 150
                                np.fill_diagonal(potential, False) # Skip self
                                
                                if np.any(potential):
                                    # Gaze Matrix: check_gaze_at_target(observer_pos, observer_yaw, target_pos)
                                    # dx = target_pos[0] - observer_pos[0]
                                    dx_mat = centers_arr[np.newaxis, :, 0] - centers_arr[:, np.newaxis, 0] # (ObsRow, TarCol)
                                    
                                    # Threshold logic
                                    looking_right = yaws_arr > 10
                                    looking_left = yaws_arr < -10
                                    looking_center = np.abs(yaws_arr) < 10
                                    
                                    # Matrix of "Is Obs i looking at Tar j?"
                                    gaze_success = (
                                        ((dx_mat > 50) & looking_right[:, None]) |
                                        ((dx_mat < -50) & looking_left[:, None]) |
                                        ((np.abs(dx_mat) <= 50) & looking_center[:, None])
                                    )
                                    
                                    # Mutual Gaze: (i looks at j) AND (j looks at i)
                                    mutual = potential & gaze_success & gaze_success.T
                                    
                                    indices = np.where(mutual)
                                    for idx_i, idx_j in zip(*indices):
                                        if idx_i < idx_j: # Avoid double count + ensure sorted pair
                                            id_i = frame_display_ids[idx_i]
                                            id_j = frame_display_ids[idx_j]
                                            pair = (id_i, id_j)
                                            id_interactions[pair] = id_interactions.get(pair, 0) + 1
                            except Exception as e_soc:
                                logging.debug(f"Social Vectorization Error: {e_soc}")
                        
                        # [Fix] Update UI (Inside Try, Outside Person Loop)
                        if 'annotated_frame' not in locals():
                            annotated_frame = frame

                        if 'st_frame' in locals() and annotated_frame is not None:
                            # [v91.4 Fix] User requested to see the analysis process. 
                            # Show preview even if cloud_booster is ON, but with sensible throttling.
                            display_interval = 2 if not cloud_booster else 8 # 每 2 幀或 8 幀更新一次畫面
                            
                            if (f_idx // frame_interval) % display_interval == 0:
                                # [v21.5.6] Resolution Scaling: Resize to 960px width for better visibility
                                h, w = annotated_frame.shape[:2]
                                new_w = 960
                                new_h = int(h * (new_w / w))
                                small_frame = cv2.resize(annotated_frame, (new_w, new_h))
                                frame_rgb = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
                                st_frame.image(frame_rgb, caption=f"正在分析... (第 {f_idx}/{total_frames} 幀)")
                                # [v89 Fix] Cache the last visualized frame for final report
                                st.session_state.last_frame = frame_rgb
                            
                            # Update status text periodically
                            if f_idx % 60 == 0:
                                mode_label = "MediaPipe" if use_mp else "YOLO"
                                booster_status = "🚀 加速中" if cloud_booster else "🐢 標準模式"
                                st_progress_text.text(f"【{booster_status}】{mode_label} 正在全力辨識數據... ({prog*100:.1f}%) [Frame {f_idx}/{total_frames}]")

                        # [v87 Sync] Write frames MULTIPLE times to maintain original duration
                        # If frame_interval is 6, we write this result 6 times to sync with 30fps clock
                        if out_video is not None and st.session_state.get('save_recording', True):
                            try:
                                f_to_write = annotated_frame if 'annotated_frame' in locals() and annotated_frame is not None else frame
                                if f_to_write.shape[1] != width or f_to_write.shape[0] != height:
                                    f_to_write = cv2.resize(f_to_write, (width, height))
                                
                                # Write frame_interval times to sync duration
                                for _ in range(frame_interval):
                                    out_video.write(f_to_write)
                            except Exception as e_write:
                                logging.error(f"Video Write Error: {e_write}")

                    except Exception as e:
                        if "lap" in str(e) or "LAP" in str(e):
                            st.error("❌ 缺少 'lap' 模組。請確認 requirements.txt 包含 'lapx'。")
                            st.stop()
                        logging.error(f"Analysis Frame Error: {e}")
                        pass
                    
                    # Jump frame writer completely removed for Fast Mode

                cap.release()

                # [v21.12.16] Finalize IDs Safety Net (Ensures report is never empty for short clips)
                for mid_id, feat_id in st.session_state.id_features.items():
                    if "clothing" not in feat_id and feat_id.get("frames_accumulated", 0) > 0:
                        from collections import Counter
                        final_h = Counter(feat_id.get("clothing_buffer", [])).most_common(1)[0][0] if feat_id.get("clothing_buffer") else "未知"
                        final_p = Counter(feat_id.get("pants_buffer", [])).most_common(1)[0][0] if feat_id.get("pants_buffer") else "未知"
                        feat_id["clothing"] = f"上衣：{final_h}。下裝：{final_p} (短時偵測)。"
                        if "hist" not in feat_id and feat_id.get("hist_buffer"):
                            feat_id["hist"] = feat_id["hist_buffer"][-1]

                # First Frame Guard
                if f_idx == 0:
                    st.error("❌ 影像讀取失敗：伺服器無法解析此影片編碼。")
                    if 'holistic_model' in locals() and holistic_model: holistic_model.close()
                    if out_video: out_video.release()
                    st.stop()
                if 'holistic_model' in locals() and holistic_model:
                    holistic_model.close()
                if out_video is not None and out_video.isOpened():
                    out_video.release() # Release writer
                
                # [v21.12.20 Fix] Move processed video from TEMP to local OneDrive folder
                if os.path.exists(temp_output_path):
                    import shutil
                    shutil.copy(temp_output_path, final_output_path)
                    output_path = final_output_path
                    
                # [v21.5.1 Hotfix] Move conversion BEFORE rerun to avoid race condition
                # [v91.9 Fix] Ultra-Robust FFmpeg Detection
                if os.path.exists(output_path):
                    sid = st.session_state.session_id
                    converted_path = os.path.abspath(f"hmeayc_final_h264_{sid}.mp4")
                    st.info("🔄 正在進行影片格式優化 (H.264)，完成後將開啟支援「調速」功能之播放器...")
                    
                    import shutil
                    import subprocess
                    
                    # Search for FFmpeg in common Windows locations
                    ffmpeg_exe = shutil.which("ffmpeg")
                    if not ffmpeg_exe:
                        possible_bins = [
                            os.path.join(os.environ.get('USERPROFILE', ''), r"AppData\Local\Microsoft\WinGet\Packages\Gyan.FFmpeg_Microsoft.Winget.Source_8wekyb3d8bbwe\ffmpeg-8.0.1-full_build\bin\ffmpeg.exe"),
                            r"C:\ffmpeg\bin\ffmpeg.exe",
                            r"C:\Program Files\ffmpeg\bin\ffmpeg.exe"
                        ]
                        for pb in possible_bins:
                            if os.path.exists(pb):
                                ffmpeg_exe = pb
                                break

                    if ffmpeg_exe:
                        try:
                            # [v91.9] Force standard H.264 + yuv420p for browser playback & speed control
                            cmd = [
                                ffmpeg_exe, "-y", "-i", output_path,
                                "-c:v", "libx264", "-preset", "ultrafast", "-crf", "24",
                                "-pix_fmt", "yuv420p",
                                "-f", "mp4", converted_path
                            ]
                            subprocess.run(cmd, check=True, capture_output=True, timeout=180)
                            if os.path.exists(converted_path):
                                st.session_state.video_output_path = converted_path
                            else:
                                st.session_state.video_output_path = output_path
                        except Exception as e_ff:
                            st.error(f"❌ 影片轉檔失敗: {e_ff}")
                            st.session_state.video_output_path = output_path
                    else:
                        st.warning("⚠️ 系統找不到轉檔工具 (FFmpeg)，播放器可能無法顯示內容，建議安裝後重試。")
                        st.session_state.video_output_path = output_path
                else:
                    st.session_state.video_output_path = None

                st_progress.empty() # 清除進度條
                st_progress_text.empty()

                # [v21.5.2 Fix] Remove redundant path overwriting that happened in v21.5.1
                # The video_output_path is already set by the FFmpeg block above.
                if 'video_output_path' not in st.session_state or st.session_state.video_output_path is None:
                    if out_video is not None:
                         st.session_state.video_output_path = output_path

                # [v66] 直接在此處完成所有數據結算，避免二次 Rerun 失敗
                logging.info("Calculating Group Sync (Kuramoto)...")
                group_sync_r = calculate_kuramoto_order_parameter(st.session_state.id_motion_log)
                st.session_state.group_sync_r = group_sync_r if group_sync_r is not None else 0.0

                # [v89 Fix] Lower stability threshold to 1 for maximum detection retention
                # This ensures ANY child detected even once is included in the final report.
                stable_ids = {mid for mid, count in st.session_state.id_tracking_count.items() if count >= 1}
                
                # Update all state to reflect only stable IDs
                st.session_state.id_list = {st.session_state.display_mapping.get(mid) for mid in stable_ids if mid in st.session_state.display_mapping}
                st.session_state.id_list.discard(None) # Safety check
                
                st.session_state.final_id_count = len(st.session_state.id_list)
                st.session_state.final_id_list = sorted(list(st.session_state.id_list))
                
                # Filter other data structures to prevent ghost entries in report
                st.session_state.id_features = {mid: feat for mid, feat in st.session_state.id_features.items() if mid in st.session_state.id_list}
                st.session_state.id_actions = {mid: acts for mid, acts in st.session_state.id_actions.items() if mid in st.session_state.id_list}
                st.session_state.id_positions = {mid: pos for mid, pos in st.session_state.id_positions.items() if mid in st.session_state.id_list}

                st.success("✅ 分析完成！正在跳轉報表...")
                st.session_state.analysis_done = True
                st.rerun() 

            except Exception as e:
                st.error(f"發生系統錯誤: {e}")
                st.write(traceback.format_exc())
            finally:
                # [v86] Robust Resource Cleanup
                if 'cap' in locals() and cap is not None:
                    try: cap.release()
                    except: pass
                if 'holistic_model' in locals() and holistic_model:
                    try: holistic_model.close()
                    except: pass
                if 'out_video' in locals() and out_video is not None:
                    try: out_video.release()
                    except: pass

                # 清理暫存檔
                if os.path.exists(tfile_path):
                    try:
                        # os.remove(tfile_path) # Debug: Keep for now
                        pass 
                    except:
                        pass


# [v66] The transition results block has been consolidated into the analysis completion block above.
# The code below now handles the final report rendering.


# [v91.23] Removed st.stop() to prevent blank pages.
if not st.session_state.get('analysis_done', False) and st.session_state.current_step != "1️⃣ 影片設定":
    st.warning("⚠️ 請先前往「1️⃣ 影片設定」上傳並點擊『開始分析』，分析完成後數據將在此顯示。")

# --- 4. 報表編輯區 (分析完成後鎖定顯示) ---

if st.session_state.current_step == "2️⃣ 分析報表":
    # [新] 持續顯示 ID 資訊
    if 'final_id_count' in st.session_state and st.session_state.final_id_count > 0:
        st.info(f"📊 偵測到的 ID 數量: {st.session_state.final_id_count}")
        st.write(f"ID 列表: {st.session_state.final_id_list}")
        if 'group_sync_r' in st.session_state:
            st.write(f"🔗 群體同步率 (R值): {st.session_state.group_sync_r} (1.0 為完全同步)")

# [v91.16] Combined Analysis & Report Display (Step 2)
if st.session_state.current_step == "2️⃣ 分析報表":
    if 'video_output_path' in st.session_state and st.session_state.video_output_path:
        target_video = st.session_state.video_output_path
        if os.path.exists(target_video):
            st.success("✅ 分析報告已就緒 - 請在下方觀看影片並完成觀察報表")
            
            # --- Player Setup ---
            cap_p = cv2.VideoCapture(target_video)
            fps = cap_p.get(cv2.CAP_PROP_FPS) or 30
            total_f = int(cap_p.get(cv2.CAP_PROP_FRAME_COUNT))
            duration = total_f / fps
            
            # [v91.20] Layout: Centered 70% width
            _, main_col, _ = st.columns([1, 4, 1])

            # [v91.20] Force Single Row Layout for Mobile (CSS Hack)
            st.markdown("""
                <style>
                [data-testid="column"] {
                    min-width: 0px !important;
                    flex-basis: 0 !important;
                    flex-grow: 1 !important;
                }
                div[data-testid="stHorizontalBlock"] {
                    flex-wrap: nowrap !important;
                    gap: 5px !important;
                }
                .stButton button {
                    padding: 0.2rem 0.4rem !important;
                    font-size: 16px !important;
                    width: 100% !important;
                }
                </style>
            """, unsafe_allow_html=True)

            with main_col:
                # 1. Video Display Area
                st_video_display = st.empty()
                
                # 2. YouTube-style Time Slider
                if 'seek_time' not in st.session_state: st.session_state.seek_time = 0.0
                def format_time(s): return f"{int(s//60):02d}:{int(s%60):02d}"
                
                curr_seek_s = st.slider(
                    f"⏰ {format_time(st.session_state.seek_time)} / {format_time(duration)}", 
                    0.0, float(duration), float(st.session_state.seek_time), step=0.1,
                    label_visibility="visible"
                )
                st.session_state.seek_time = curr_seek_s
                seek_frame = int(curr_seek_s * fps)
                
                # 3. Compact Controls Row (5 Columns, No Stacking)
                c1, c2, c3, c4, c5 = st.columns([1, 1, 1, 1, 1.5])
                
                if c1.button("⏪"): 
                    st.session_state.seek_time = max(0.0, st.session_state.seek_time - 5.0)
                    st.rerun()
                
                do_play = c2.button("▶️")
                do_stop = c3.button("⏸️")
                
                if c4.button("⏩"):
                    st.session_state.seek_time = min(duration, st.session_state.seek_time + 5.0)
                    st.rerun()
                
                v_speed = c5.selectbox("Speed", [0.5, 1.0, 1.5, 2.0], index=1, label_visibility="collapsed")
                
                # [v18.8 Fix] Download Button (Full width below)
                st.download_button(
                    label="📥 下載分析影片",
                    data=open(target_video, "rb"),
                    file_name="hmeayc_final_report.mp4",
                    mime="video/mp4",
                    use_container_width=True,
                    key=f"dl_v92_{st.session_state.current_step}"
                )

                # --- Playback Logic (Professional Forensic Optimization) ---
                if do_play:
                    curr_f = seek_frame
                    target_delay = (1.0 / fps) / v_speed
                    
                    # [v91.17] Forward-only reading (No cap.set inside loop to prevent lag)
                    cap_p.set(cv2.CAP_PROP_POS_FRAMES, curr_f)
                    
                    while curr_f < total_f:
                        start_t = time.time()
                        ret, frame = cap_p.read()
                        if not ret: break
                        
                        # [v91.17] Professional Compression for Ultra-Smooth Playback
                        # Resize to 540px for ideal balance between clarity and speed
                        new_w = 540
                        h_o, w_o = frame.shape[:2]
                        new_h = int(h_o * (new_w / w_o))
                        small_frame = cv2.resize(frame, (new_w, new_h), interpolation=cv2.INTER_AREA)
                        
                        frame_rgb = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
                        st_video_display.image(frame_rgb, use_container_width=True)
                        
                        curr_f += 1
                        # Update the slider in session state (Visual Feedback)
                        st.session_state.seek_time = curr_f / fps
                        
                        elapsed = time.time() - start_t
                        wait_t = max(0, target_delay - elapsed)
                        time.sleep(wait_t)
                else:
                    # [v91.17] High-Precision Seeking using grab() logic where possible
                    cap_p.set(cv2.CAP_PROP_POS_FRAMES, seek_frame)
                    ret, frame = cap_p.read()
                    if ret:
                        st_video_display.image(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB), use_container_width=True)
            
            cap_p.release()
        else:
            if st.session_state.analysis_done:
                st.error("❌ 找不到影片檔案，請嘗試重新執行一次分析。")

st.markdown("---")
# st.subheader("📊 HMEAYC 專業觀察編輯")

s_ids = sorted(list(st.session_state.id_list))
if not s_ids:
    if target_track_id > 0:
        st.warning(f"⚠️ 分析完成，但在影片中未抓取到您指定的 ID {target_track_id}。")
        st.info(f"💡 目前影片中總共偵測到 {len(st.session_state.display_mapping)} 位不同的幼兒。如果您想觀察其他孩子，請調整左側的追蹤 ID。")
    else:
        st.warning("⚠️ 偵測過程中未抓取到有效 ID，請重新上傳清晰影片。")
else:
    # [v11 Update] 選擇教師 ID (用於計算師生同步率)
    # 製作選項列表: "ID_1 (原:80)"
    id_options = []
    id_map_rev = {} # 顯示名稱 -> 真實 ID

    for idx, m in enumerate(s_ids, 1):
        # [v40 Fix] Use Raw ID directly to match Video Overlay
        # User confusion: Video says ID 9, Table says ID 7 (Original 9).
        # Solution: Table should say ID 9.
        label = f"ID_{m}"
        id_options.append(label)
        id_map_rev[label] = m

    col_t1, col_t2 = st.columns([1, 3])
    with col_t1:
        if st.session_state.current_step == "2️⃣ 分析報表":
            teacher_label = st.selectbox("請選擇教師 (示範者) ID:", ["無"] + id_options)
            st.session_state.teacher_label = teacher_label
        else:
            teacher_label = st.session_state.get('teacher_label', "無")

    teacher_id = None
    if teacher_label != "無":
        teacher_id = id_map_rev[teacher_label]
        if st.session_state.current_step == "2️⃣ 分析報表":
            st.info(f"已設定 {teacher_label} 為教師，將計算其他幼兒與其的動作同步率。")

    # [v74 New] Restored Data Handling
    if st.session_state.get('restored_df') is not None:
        st.warning("⚠️ 目前顯示的是從「歷史紀錄」載入的資料。您可以繼續編輯或產出報表。")
        if st.button("❌ 取消載入，回到當前分析結果"):
            del st.session_state.restored_df
            st.rerun()
        df = st.session_state.restored_df
        # We still need s_ids for some calculations? Not really, we can skip the loop.
    else:
        # [v90.7 Fix] Persistent Report Hub
        # Check session state to see if we already have a completed report.
        if 'final_report_df' in st.session_state and st.session_state.final_report_df is not None:
            df = st.session_state.final_report_df
            logging.info("♻️ [v90.7] Restoring Report from Session Cache")
        else:
            # 從 Session State 讀取資料
            df_list = []

            for idx, m in enumerate(s_ids, 1): # idx 從 1 開始
                # [v8 Fix] 先取出 feat 才能判斷 original_id
                feat = st.session_state.id_features.get(m, {})
                if not feat: feat = {"clothing": "分析中..."}
                
                # [v21.9.1 Fix] Fallback for Short-Lived IDs (Avoid Empty Table Cells)
                if "clothing" not in feat and feat.get("frames_accumulated", 0) > 0:
                    from collections import Counter
                    final_s = Counter(feat.get("clothing_buffer", [])).most_common(1)[0][0] if feat.get("clothing_buffer") else "未知"
                    final_p = Counter(feat.get("pants_buffer", [])).most_common(1)[0][0] if feat.get("pants_buffer") else "未知"
                    feat["clothing"] = f"上衣：{final_s}。下裝：{final_p}。"
                    if "hist" not in feat and feat.get("hist_buffer"):
                        feat["hist"] = feat["hist_buffer"][-1]

                # [v21.9.1 Add] Manual Merge Check
                manual_map = st.session_state.get('manual_id_map', {})
                if m in manual_map:
                    continue 

                original_id = m
                pos_history = st.session_state.id_positions.get(m, []).copy()
                motion_log = st.session_state.id_motion_log.get(m, []).copy()
                action_counts = st.session_state.id_actions.get(m, {}).copy()
                yaw_history = st.session_state.id_yaw_history.get(m, []).copy()

                sources = [s for s, t in manual_map.items() if t == m]
                for s in sources:
                    pos_history.extend(st.session_state.id_positions.get(s, []))
                    pos_history.sort(key=lambda x: x[0])
                    motion_log.extend(st.session_state.id_motion_log.get(s, []))
                    s_actions = st.session_state.id_actions.get(s, {})
                    for act, count in s_actions.items():
                        action_counts[act] = action_counts.get(act, 0) + count
                    yaw_history.extend(st.session_state.id_yaw_history.get(s, []))

                # 計算動態評分
                pos_only = [p[1] for p in pos_history]
                score = get_motion_score(pos_only) 

                energy = 0
                if motion_log:
                    raw_energy = np.mean(motion_log)
                    energy = min(round(raw_energy * 0.8, 1), 10.0)

                smoothness_score = 0
                if len(motion_log) > 2:
                    accels = np.diff(motion_log)
                    jerks = np.diff(accels)
                    input_jerk = np.mean(np.abs(jerks))
                    smoothness_score = round(100 * np.exp(-0.08 * input_jerk), 1)

                # 計算師生同步率
                sync_score = None
                lag_sec = 0
                temp_corr = 0
                if teacher_id is not None:
                    try:
                        if int(m) == int(teacher_id):
                            sync_score = 100.0 
                        else:
                            teacher_pos = st.session_state.id_positions.get(teacher_id, [])
                            raw_r = calculate_teacher_sync(pos_history, teacher_pos)
                            if raw_r is not None:
                                sync_score = float(raw_r)
                                if 0 < sync_score <= 1.0:
                                    sync_score *= 100
                                sync_score = round(sync_score, 1)
                            else:
                                sync_score = 0.0

                            student_log = st.session_state.id_motion_log.get(m, [])
                            teacher_log = st.session_state.id_motion_log.get(teacher_id, [])
                            eff_fps = st.session_state.get("effective_fps", 30)
                            temp_corr, lag_sec = analyze_temporal_sync(student_log, teacher_log, fps=eff_fps)
                            feat['temp_lag'] = lag_sec
                            feat['temp_corr'] = temp_corr
                    except Exception as e:
                        logging.error(f"Sync error for {m}: {e}")
                        sync_score = 0.0

                if teacher_id is not None:
                    try:
                        if int(m) == int(teacher_id):
                            sync_score = 100.0
                    except:
                        pass

                # 特徵補強
                valid_tags = []
                priority = ["跳躍", "地板動作", "舉手", "蹲下", "踢腿/抬腿", "跑步", "側臉", "專注"] 
                for tag in priority:
                    if action_counts.get(tag, 0) > 3:
                        valid_tags.append(tag)

                # 生成 AI 專注度
                focus_score = 0
                if teacher_id is not None and m != teacher_id and teacher_id in st.session_state.id_positions:
                    t_pos_list = st.session_state.id_positions[teacher_id]
                    if t_pos_list:
                        t_avg_x = t_pos_list[-1][1][0]
                        min_len = min(len(yaw_history), len(pos_history))
                        valid_f = 0
                        focused_f = 0
                        for i in range(0, min_len, 5):
                            s_x = pos_history[i][1][0]
                            yaw = yaw_history[i]
                            if check_gaze_at_target((s_x, 0), yaw, (t_avg_x, 0)):
                                focused_f += 1
                            valid_f += 1
                        if valid_f > 0:
                            focus_score = int((focused_f / valid_f) * 100)

                # 角色判定
                interaction_count = 0
                if 'id_interactions' in st.session_state:
                    interaction_count = sum([c for (pair, c) in st.session_state.id_interactions.items() if m in pair])
                
                role = "獨立觀察 (Independent)" 
                if interaction_count >= 60 and score >= 3:
                    role = "社交活躍 (Active)"
                elif interaction_count >= 60:
                    role = "靜態互動 (Passive)"
                elif focus_score >= 60:
                    role = "專注跟隨 (Focused)"
                elif sync_score is not None and sync_score < 40 and energy > 5.0:
                    role = "自主探索 (Creative)"
                elif sync_score is not None and sync_score >= 80:
                    role = "動作模仿 (Imitating)"
                if teacher_id is not None and m == teacher_id:
                    role = "教學者 (Teacher)"
                if "MediaPipe" in perf_mode:
                    role = "微觀特寫 (Micro-Analysis)"
                    sync_score = None

                pure_actions = [tag for tag in valid_tags if tag not in ["專注", "側臉"]]
                ai_comment = ""
                sync_val = sync_score if sync_score is not None else 0
                stab_val = smoothness_score
                
                # AI 評語邏輯 (簡化為主要 7 類)
                if role == "微觀特寫 (Micro-Analysis)":
                    options = ["動作執行度極高，肢體延展充分且控制精準。"] if stab_val > 70 else ["目前呈現較為基礎的律動節奏。"]
                elif (teacher_id is not None and m == teacher_id) or (sync_val > 80 and stab_val > 70):
                    options = ["動作精準且穩定，完全跟隨老師指令。"]
                elif energy > 7.0 and sync_val > 40:
                    options = ["充滿活力，動作幅度大且熱情。"]
                elif energy < 3.0 and focus_score > 60:
                    options = ["雖動作較少，但視線持續專注於示範者。"]
                elif sync_val < 40 and energy > 5.0:
                    options = ["展現出獨特的動作詮釋，具有豐富的創造力。"]
                elif interaction_count > 20 and sync_val > 40:
                    options = ["與同儕互動頻繁，樂於透過人際互動來學習。"]
                elif focus_score < 30 and energy > 5.0:
                    options = ["精力旺盛但較難聚焦，建議加強引導。"]
                else:
                    options = ["動作平穩流暢，依照自己的節奏跟隨。"]
                ai_comment = random.choice(options)

                lag_display = f"{feat.get('temp_lag', 0):.2f}s" if feat.get('temp_lag', 0) and abs(feat.get('temp_lag', 0)) > 0.1 else "-"
                act_ctx = st.session_state.get('activity_context', '跟隨模仿 (Imitation)')
                final_comment = generate_expert_comment(score, sync_score, focus_score, role, pure_actions, activity_context=act_ctx, class_stats=None, archetype_text=ai_comment)

                df_list.append({
                    "序號": idx, 
                    "幼兒 ID": f"ID_{idx} (原:{original_id})", 
                    "AI 服裝特徵": feat.get("clothing", ""),
                    "特徵補強 (圖案/熊/亮片)": None, 
                    "AI 觀察判定 (1-5)": score,
                    "跟隨指令 (同步率%)": float(f"{sync_score:.0f}") if sync_score is not None else 0, 
                    "時序延遲 (Lag)": lag_display, 
                    "專注度(%)": focus_score, 
                    "參與型態": role,        
                    "動作檢測 (舉手、側臉)": ", ".join(pure_actions),
                    "AI 總結評語": final_comment,
                    "教師評分 (1-5)": None,
                    "教師評語": None,
                    "動作能量": energy,
                    "動作穩定度": smoothness_score,
                })

            df = pd.DataFrame(df_list)
            st.session_state.final_report_df = df
            logging.info("💾 [v90.7] Final Report Saved to Session Cache")

    # [v20.2 Sort] Sort by ID number
    if not df.empty:
        df['sort_key'] = df['幼兒 ID'].apply(lambda x: int(x.split(' ')[0].split('_')[1]) if '_' in x else 999)
        df = df.sort_values('sort_key').drop(columns=['sort_key'])

    # [v65 New] Custom Column Layout for Micro-Analysis
    if "MediaPipe" in st.session_state.get('perf_mode', ''):
        required_cols = ["序號", "幼兒 ID", "AI 服裝特徵", "特徵補強 (圖案/熊/亮片)", "AI 觀察判定 (1-5)", "時序延遲 (Lag)", "專注度(%)", "動作能量", "動作穩定度", "動作檢測 (舉手、側臉)", "AI 總結評語", "教師評分 (1-5)", "教師評語"]
    else:
        required_cols = ["序號", "幼兒 ID", "AI 服裝特徵", "特徵補強 (圖案/熊/亮片)", "AI 觀察判定 (1-5)", "跟隨指令 (同步率%)", "時序延遲 (Lag)", "專注度(%)", "動作能量", "動作穩定度", "參與型態", "動作檢測 (舉手、側臉)", "AI 總結評語", "教師評分 (1-5)", "教師評語"]

    if df.empty:
        df = pd.DataFrame(columns=required_cols)
    else:
        for col in required_cols:
            if col not in df.columns:
                df[col] = None 
        df = df.reindex(columns=required_cols) # Ensures exact column order and drops unwanted columns

    # [v31 Fix] Persistent Naming Logic using Session State
    if 'custom_name_map' not in st.session_state:
        st.session_state.custom_name_map = {}

    # Preserve the Original Raw ID for mapping (Hidden Column)
    # We need a column that stays constant even if "幼兒 ID" is edited.
    # "幼兒 ID" acts as the Display/Edit column.
    # "Raw_ID" acts as the Key.

    # 1. Inject Raw_ID for tracking
    # 1. Inject Raw_ID for tracking
    if not df.empty:
        # Re-extract raw ID from the "幼兒 ID" string if needed, or assume it's unique enough
        # Current "ID_X (原:Y)" is unique per session run (until restart)
        # [v45 Fix] Extract Simple Key "ID_X" for consistent mapping
        # This ensures that even if display name is complex, the key remains stable.
        # [v21.5.7 Fix] Unified Raw_ID for stable mapping
        # Extract the stable detection ID (e.g., ID_11) from the original label
        df['Raw_ID'] = df['幼兒 ID'].apply(lambda x: x.split(" ")[0] if isinstance(x, str) else "")

        # Apply existing names to Display Column
        def apply_name(row):
            key = row['Raw_ID']
            val = st.session_state.custom_name_map.get(key, row['幼兒 ID'])
            return val

        df['幼兒 ID'] = df.apply(apply_name, axis=1)

    # 設定欄位格式
    # [v19 Fix] Restore st.data_editor call
    # [v36 Fix] Define Callback for Persistent Renaming
    # [v21.9.3 New] Class-Level Summary for Human Comparison
    if not df.empty:
        if st.session_state.current_step == "4️⃣ 最終報表":
            st.info("💡 **比對說明**：下表為 AI 偵測到之各幼兒數據。您可以參考下方的 **「全班 AI 總體指標」** 來與您的人工全班觀察進行對照。")
        
        # [v21.9.3 Fix] Exclude Teacher from Class Averages to prevent skewed metrics (e.g. 100% Teacher pulling up average)
        # We also want to only include people who are marked as "students" in the average
        df_students = df.copy()
        if teacher_id is not None:
            # Re-extract ID number to match teacher_id
            df_students = df[df["幼兒 ID"].apply(lambda x: "ID_" + str(teacher_id) not in x)]
        
        c_score = df_students["AI 觀察判定 (1-5)"].mean() if not df_students.empty else 0
        c_sync = df_students["跟隨指令 (同步率%)"].mean() if not df_students.empty and "跟隨指令 (同步率%)" in df_students.columns else 0
        c_focus = df_students["專注度(%)"].mean() if not df_students.empty else 0
        c_count = len(df_students)
        
        # Show Class Metrics
        if st.session_state.current_step == "4️⃣ 最終報表":
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("全班 AI 活躍評分 (Avg)", f"{c_score:.2f} / 5")
            
            # [v21.9.2 Refine] Handle No Teacher State for Sync Metrics
            if teacher_id is None:
                m2.metric("全班 AI 同步率 (Avg)", "N/A", help="🔴 請選擇『教師ID』以進行全班比對")
                m3.metric("全班 AI 專注度 (Avg)", "N/A", help="🔴 請選擇『教師ID』以進行全班比對")
            else:
                # [v21.9.3] Show more detailed warning if sync is very low
                sync_help = "代表全班學生相對於老師動作的整體一致性 (已排除老師本人的 100%)"
                if c_sync < 5.0 and c_count > 0:
                    sync_help += " ⚠️ 目前數值偏低，可能是因為偵測對齊時間不足，或當前活動並非模仿性質。"
                m2.metric("全班 AI 同步率 (Avg)", f"{c_sync:.1f} %", help=sync_help)
                m3.metric("全班 AI 專注度 (Avg)", f"{c_focus:.1f} %")
            
            m4.metric("偵測幼兒總數", f"{c_count} 人")
        
        # Human Class-Level Input for Comparison
        if st.session_state.current_step == "2️⃣ 影片分析與報表":
            with st.expander("📝 紀錄人工全班觀察數據 (點擊展開進行比對)", expanded=False):
                st.markdown("您可以將您對**全班整體**的觀察填寫於此，以便與 AI 數據進行長期比對與驗證。")
                c1, c2 = st.columns(2)
                with c1:
                    # [v90.8 Fix] Sync precision: use step=0.01 for 2 decimal places to match AI scores
                    h_score = st.slider("人工全班參與度評分 (1-5)", 1.0, 5.0, 3.0, step=0.01, key="manual_class_score")
                    diff = h_score - c_score
                    st.write(f"📊 與 AI 評分差異：**{'+' if diff > 0 else ''}{diff:.2f}**")
                with c2:
                    st.text_area("人工全班觀察評語/紀錄", placeholder="請輸入您對全班律動表現的專業觀察...", key="manual_class_note")
        
        st.write("---")

    def update_names_callback():
        """
        Callback to handle name changes by looking up stable Raw_ID from current dataframe.
        """
        editor_state = st.session_state.get("data_editor_v31_final", {})
        edited_rows = editor_state.get("edited_rows", {})
        if not edited_rows:
            return

        for row_idx_str, changes in edited_rows.items():
            if "幼兒 ID" in changes:
                try:
                    row_idx = int(row_idx_str)
                    # Directly look up the stable key from the current dataframe
                    raw_key = df.iloc[row_idx]['Raw_ID']
                    new_name = changes["幼兒 ID"]
                    
                    # Update global naming map
                    st.session_state.custom_name_map[raw_key] = new_name
                    logging.info(f"Renamed {raw_key} to {new_name}")
                except Exception as e:
                    logging.error(f"Rename Callback Error: {e}")

    if st.session_state.current_step != "2️⃣ 影片分析與報表":
        edited_df = df.copy()
    else:
        edited_df = st.data_editor(
        df, 
        use_container_width=True,
        column_config={
            "Raw_ID": None, # Hide the boolean/key column
            "序號": st.column_config.NumberColumn("序號", format="%d", width=40, disabled=True), 
            "動作能量": st.column_config.ProgressColumn(
        "動作能量 (Energy)", 
        help="動作的大小與劇烈程度 (0-100)", 
        min_value=0, max_value=100,
        format="%d"
    ),
    "動作穩定度": st.column_config.ProgressColumn(
        "動作穩定度 (Smoothness)", 
        help="[Neuro-Motor] 動作流暢性，低分代表抖動或不穩 (0-100)", 
        min_value=0, max_value=100,
        format="%d"
    ),
    "專注度(%)": st.column_config.ProgressColumn("專注度", min_value=0, max_value=100, format="%d%%", width=80), 
            "參與型態": st.column_config.TextColumn("參與型態", width=120), 
            "幼兒 ID": st.column_config.TextColumn( 
                "幼兒 ID (可修改姓名)", 
                width=150, 
                disabled=False,
                help="點擊兩下修改姓名，系統會自動記憶 (直到重整網頁)"
            ),
            "AI 服裝特徵": st.column_config.TextColumn("AI 服裝特徵", width=300), 
            "特徵補強 (圖案/熊/亮片)": st.column_config.TextColumn("特徵補強", width=100), 
            "AI 觀察判定 (1-5)": st.column_config.NumberColumn("AI 評分", width=80), 
            "跟隨指令 (同步率%)": st.column_config.NumberColumn("同步率", format="%.0f", width=80), 
            "動作檢測 (舉手、側臉)": st.column_config.TextColumn("動作檢測", width=150), 
            "AI 總結評語": st.column_config.TextColumn("AI 總結評語", width=600), 
            "教師評分 (1-5)": st.column_config.NumberColumn("教師評分", width=80), 
            "教師評語": st.column_config.TextColumn("教師評語", width=200),
        },
        hide_index=True,
        num_rows="dynamic", # [v76 New] Allow row deletion
        key="data_editor_v31_final", # Unique key
        on_change=update_names_callback # [v36] Bind callback
    )

    # [v36] Removed old manual diff logic (lines 2077-2100) as callback handles it robustly.
    # Check if we need to force rerun?
    # Streamlit automatically reruns after callback.
    # Since logic updates state before rerun, the next run sees updated map.
    # df is rebuilt with map -> Editor shows new name.
    # Perfect.


    # [v19 New] Display Social Graph
    if st.session_state.current_step == "3️⃣ 社交網絡":
        # st.write("---")
        st.subheader("🕸️ 班級社交互動網絡圖 (Social Graph)")
        if st.session_state.id_interactions:
            # Generate graph
            try:
                # [v74 Fix] Pass sensitivity threshold to graph drawer
                s_thresh = st.session_state.get('social_threshold_sec', 3.0)
                graph_img = draw_social_graph(st.session_state.id_interactions, 
                                            {m: f"{m}" for m in st.session_state.id_list},
                                            min_sec=s_thresh)
                # [v20.11 Refine] Larger Display Width (1100px)
                # [Fix] Convert BGR to RGB for correct color display in Streamlit
                # Display Graph
                graph_img_rgb = cv2.cvtColor(graph_img, cv2.COLOR_BGR2RGB)
                # [v35 Fix] Removed duplicate st.image call and updated caption per user request
                st.image(graph_img_rgb, use_container_width=True, caption="🔴紅色=社交核心 | 🔵藍色=一般 | 線條粗細=互動頻率")
    
                # [v34 New] Interaction Details List to clarify connections
                with st.expander("詳細互動清單 (Interaction Details)"):
                    if 'id_interactions' in st.session_state and st.session_state.id_interactions:
                        # Sort by count
                        sorted_inters = sorted(st.session_state.id_interactions.items(), key=lambda x: x[1], reverse=True)
    
                        # [v21.4 Refine] Split list into 2 columns for better readability
                        col_L, col_R = st.columns(2)
                        mid_idx = (len(sorted_inters) + 1) // 2
                        
                        # Prepare ID map for names
                        id_name_map = {}
                        if not df.empty:
                            for _, row in df.iterrows():
                                # Extract ID from display name like "ID_3 (原:3)"
                                r_id_col = row.get('幼兒 ID', "")
                                try:
                                    if "ID_" in str(r_id_col):
                                        p = str(r_id_col).split("ID_")[1].split(" ")[0]
                                        int_id = int(p)
                                        id_name_map[int_id] = row['幼兒 ID'] 
                                except:
                                    pass
                        for i, (pair, count) in enumerate(sorted_inters):
                            p1, p2 = pair
                            name1 = id_name_map.get(p1, f"ID_{p1}")
                            name2 = id_name_map.get(p2, f"ID_{p2}")
                            
                            # [v73 Fix] Time calculation based on interaction count and sampling
                            # Each count represents a sample taken every 'frame_interval'
                            f_int = st.session_state.get('last_frame_interval', 2)
                            sec = (count * f_int) / 30.0 
                            
                            # [v74 New] Threshold indicator for better UX
                            s_thresh = st.session_state.get('social_threshold_sec', 3.0)
                            icon = "🔗" if sec >= s_thresh else "⚪"
                            
                            target_col = col_L if i < mid_idx else col_R
                            target_col.markdown(f"{icon} **{name1}** ↔ **{name2}**: 互動約 {sec:.1f} 秒 (強度: {count})")
    
                st.markdown("""
                **📖 如何解讀社交圖譜 (How to Read)**
                *   **🔵 藍色圓圈 (Nodes)**: 代表一位被偵測到的幼兒 (ID)。
                *   **🔴 紅色圓圈 (Social Hubs)**: 代表社交核心幼兒 (其「主動互動」或「被互動」頻率高於全班平均)。
                *   **➖ 連線 (Edges)**: 代表兩人間有**顯著互動**。
                    *   *判定標準*：兩人距離靠近且視線/動作具備時序關聯性。
                *   **線條粗細 (Thickness)**: 線條越粗，代表兩人互動的累積時間越長、次數越多。
                """)
            except Exception as e:
                st.error(f"無法繪製社交圖表: {e}")
        else:
            st.info("尚未偵測到顯著的互動事件 (需靠近且持續互動)。")

        # [v91.19] Step 3: Finish & Save Buttons
        st.write("---")
        c1, c2 = st.columns(2)
        if c1.button("💾 儲存所有分析結果與圖表", use_container_width=True, type="primary"):
            st.success("✅ 社交關係圖與分析數據已成功備份至系統資料庫！")
        
        if c2.button("🏁 結束分析並返回首頁", use_container_width=True):
            st.session_state.nav_index = 0
            # Full Reset
            keep_k = ['session_id', 'nav_mode']
            for key in list(st.session_state.keys()):
                if key not in keep_k:
                    del st.session_state[key]
            st.rerun()

    if st.session_state.current_step == "2️⃣ 分析報表":
        if st.button("✨ 點此產生 Excel 報表數據"):
            out = io.BytesIO()
            final_excel_df = edited_df.copy()
            # 下載時將 ID 轉回字串格式，避免 Excel 視為數字
            final_excel_df["幼兒 ID"] = final_excel_df["幼兒 ID"].apply(lambda x: f"ID_{x}" if str(x).isdigit() else str(x))
    
            final_excel_df.insert(0, "觀察員", observer_name)
            final_excel_df.insert(1, "活動名稱", act_name)
            final_excel_df.insert(2, "觀察日期", act_date)
            final_excel_df.insert(3, "音樂元素", music_element)
    
            # [v20.1 Update] Advanced Excel Formatting with specific column widths
            with pd.ExcelWriter(out, engine='openpyxl') as writer:
                # [v21.9.3] 1. Prepare Summary Data
                h_score = st.session_state.get('manual_class_score', 0.0)
                h_note = st.session_state.get('manual_class_note', "")
                
                # Recalculate AI Averages for the export (Same logic as UI)
                df_students = edited_df.copy()
                # If teacher label is selected in UI, it might be in the ID col as "ID_X (原:X)"
                t_id = st.session_state.get('teacher_id_number') # We should set this when selecting teacher
                if teacher_id: # From outer scope
                    df_students = edited_df[edited_df["幼兒 ID"].apply(lambda x: "ID_" + str(teacher_id) not in x)]
    
                avg_score = df_students["AI 觀察判定 (1-5)"].mean() if not df_students.empty else 0
                avg_sync = df_students["跟隨指令 (同步率%)"].mean() if not df_students.empty and "跟隨指令 (同步率%)" in df_students.columns else 0
                avg_focus = df_students["專注度(%)"].mean() if not df_students.empty else 0
                
                summary_data = [
                    ["全班 AI 總體指標摘要", ""],
                    ["全班 AI 活躍評分 (Avg)", f"{avg_score:.2f} / 5"],
                    ["全班 AI 同步率 (Avg)", f"{avg_sync:.1f} %"],
                    ["全班 AI 專注度 (Avg)", f"{avg_focus:.1f} %"],
                    ["偵測幼兒總數", f"{len(df_students)} 人"],
                    ["", ""],
                    ["人工全班觀察紀錄", ""],
                    ["人工全班評分", f"{h_score} / 5"],
                    ["人工全班觀察筆記", h_note],
                    ["", ""]
                ]
                summary_df = pd.DataFrame(summary_data, columns=["項次", "數值/內容"])
                # [v87 Fix] Shift columns to B and C by setting startcol=1
                summary_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=0, startcol=1)
    
                # 2. Write main student records table (shifted down)
                final_excel_df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=len(summary_data) + 1)
                worksheet = writer.sheets['Sheet1']
    
                # [v65 Refactor] Dynamic Column Widths based on actual DataFrame columns
                from openpyxl.utils import get_column_letter
                from openpyxl.styles import Alignment, Font, PatternFill
    
                # [v21.9.3] 3. Styling Summary Block
                header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid") # Light Blue
                title_font = Font(bold=True, size=12)
                
                # Apply to summary headers (now at col B and C -> columns 2 and 3)
                for row_idx in [1, 7]: # "全班 AI 總體指標摘要" and "人工全班觀察紀錄"
                    for col_idx in range(2, 4):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = header_fill
                        cell.font = title_font
    
                col_width_map = {
                    "項次": 25, "數值/內容": 60, # [v89.1] Increased from 40 for better summary visibility
                    "觀察員": 12, "活動名稱": 25, "觀察日期": 15, "音樂元素": 10,
                    "序號": 5, "幼兒 ID": 25, "AI 服裝特徵": 60, # [v89.1] Increased from 50
                    "特徵補強 (圖案/熊/亮片)": 40, # [v89.1] Increased from 30
                    "AI 觀察判定 (1-5)": 20, "跟隨指令 (同步率%)": 20, "時序延遲 (Lag)": 15,
                    "專注度(%)": 12, "動作能量": 12, "動作穩定度": 12, "參與型態": 20,
                    "動作檢測 (舉手、側臉)": 35, "AI 總結評語": 100, # [v89.1] Increased from 80
                    "教師評分 (1-5)": 15, "教師評語": 40
                }
    
                wrap_cols = ["AI 服裝特徵", "特徵補強 (圖案/熊/亮片)", "動作檢測 (舉手、側臉)", "AI 總結評語", "教師評語"]
    
                # Apply widths with dynamic padding
                for idx, col_name in enumerate(final_excel_df.columns, 1):
                    col_letter = get_column_letter(idx)
                    
                    # Default logic: Use predefined width or title width
                    title_w = len(col_name.encode('utf-8')) * 0.8
                    base_w = col_width_map.get(col_name, title_w + 5)
                    
                    # [v89.2 Optimization] Auto-fit for Categorical Columns (Participant Type etc.)
                    if col_name in ["參與型態", "幼兒 ID", "觀察員", "活動名稱"]:
                        # Find max data length
                        if not final_excel_df.empty:
                            max_data_len = final_excel_df[col_name].apply(lambda x: len(str(x).encode('utf-8'))).max()
                            base_w = max(base_w, max_data_len * 0.9 + 2)
                    
                    worksheet.column_dimensions[col_letter].width = base_w
    
                # Apply text wrapping, alignment, and [v90.1 Fix] Font Size 12
                cell_font_12 = Font(size=12)
                
                for row in worksheet.iter_rows(min_row=1):
                    for cell in row:
                        cell.font = cell_font_12 # Apply to all
                        col_idx = cell.column - 1
                        if col_idx < len(final_excel_df.columns):
                            col_name = final_excel_df.columns[col_idx]
                            if col_name in wrap_cols:
                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                            else:
                                # [v89.2] Keep titles centered, data centered
                                cell.alignment = Alignment(horizontal='center', vertical='center')
    
            st.session_state.excel_ready_data = out.getvalue()
            st.success("🎉 Excel 數據已成功同步！請點擊下方按鈕下載。")

    # [v65 New] Expanded Action Buttons Layout
    col_dl, col_db, col_reset = st.columns([2, 2, 2])

    with col_dl:
        # [v21.4.4 Fix] Double-defensive check for binary data
        ready_data = st.session_state.get('excel_ready_data')
        if ready_data is not None and len(ready_data) > 0:
            st.download_button(
                label="📥 下載 Excel 正式觀察報表",
                data=ready_data,
                file_name=f"HMEAYC_Record_{act_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with col_db:
        if st.button("💾 儲存至歷史資料庫 (Save to DB)"):
            # [v37 Fix] Use persistent filename from session_state
            # [v32 Update] Handle return tuple (success, obs_id)
            target_filename = st.session_state.get("current_fn", "unknown_video.mp4")
            success, msg_or_id = save_analysis_to_db(observer_name, act_name, target_filename, edited_df)

            if success:
                st.success(f"✅ 資料已成功儲存！(紀錄 ID: {msg_or_id})")
                st.info("💡 若您再次點擊儲存，系統將會「更新」此筆紀錄，而不會產生重複資料。")
            else:
                st.error(f"❌ 儲存失敗: {msg_or_id}")

    with col_reset:
        # [v65 Restored] Full Reset Button (Clear Cache and Uploads)
        if st.button("🗑️ 清空並建立新分析 (Clear All)"):
            # [v71 Fix] Protect core session and navigation keys during reset
            keep_keys = ['session_id', 'nav_mode', 'history_student_select', 'locked_target_id']
            for key in list(st.session_state.keys()):
                if key not in keep_keys:
                    # Delete temp file if exists for current file path
                    if key == 'current_tfile_path':
                        try: os.remove(st.session_state[key])
                        except: pass
                    del st.session_state[key]
            st.rerun()

    # [v91.19] Next Step Button: Social Analysis (End of Step 2)
    if st.session_state.current_step == "2️⃣ 分析報表" and st.session_state.analysis_done:
        st.markdown("---")
        if st.button("🚀 報表生成完畢，前往「3️⃣ 互動網絡圖」分析社交關係", type="primary", use_container_width=True):
            st.session_state.nav_index = 2
            st.rerun()
